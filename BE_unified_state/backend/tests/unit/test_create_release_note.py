"""
Unit tests for create_release_note.py — comprehensive scenarios.

This file covers ALL the pure-logic functions in create_release_note.py:

1. compare_json_files()              — the top-level comparison dispatcher
2. write_changes_to_excel()          — writes change tuples to .xlsx
3. update_image_tag()                — transforms image tags between envs
4. update_image_repo_in_json_string()— transforms image tags inside JSON blobs
5. touch_helmignore()                — writes .helmignore excluding non-target envs
6. create_release_note_dir()         — creates/cleans the release note folder

Tested scenarios include:
  - Read metasheet from master branch
  - Modify / Add / Delete scenarios
  - Large YAML exceeding Excel 32K character limit
  - Pending addition scenarios
  - Pending modify scenarios
  - Hotfix scenarios (image tag transforms)
  - Root object added / deleted
  - No-change (identical configs)
"""
import json
import os
import sys
import pytest
from unittest.mock import patch, MagicMock
from openpyxl import Workbook, load_workbook

# ── Make source importable ─────────────────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'app', 'cd', 'scripts'))

from create_release_note import (
    compare_json_files,
    write_changes_to_excel,
    update_image_tag,
    update_image_repo_in_json_string,
    touch_helmignore,
    create_release_note_dir,
)
import create_release_note as crn_module
import utilities.garuda_engine as engine_module


# ═══════════════════════════════════════════════════════════════════
# Fixture: patch the module-level `envs` list
# ═══════════════════════════════════════════════════════════════════
@pytest.fixture(autouse=True)
def set_envs():
    """Patch envs in create_release_note."""
    test_envs = ['dev1', 'sit1']
    with patch.object(crn_module, 'envs', test_envs):
        yield


# ═══════════════════════════════════════════════════════════════════
# 1. compare_json_files() — TOP-LEVEL COMPARISON
# ═══════════════════════════════════════════════════════════════════

class TestCompareJsonFiles:
    """
    compare_json_files(le_old, le_new, he_old) is the entry point.
    It iterates root keys and delegates to garuda_engine.compare().

    We test it with fully constructed JSON structures that simulate
    real Helm values (service-admin.yaml, service-user.yaml, etc.).
    """

    # ── Scenario 1: No differences ─────────────────────────────
    def test_no_changes_identical_configs(self):
        """
        SCENARIO: Lower-env-old and lower-env-new are identical.
        WHAT IT TESTS: compare_json_files loops all keys but compare()
        finds no differences → returns empty list.
        REAL-WORLD: A release was cut but no config changes were made.
        """
        data = {
            "service-admin": {"replicas": 2, "image": {"image_name": "img:1.0-dev1"}},
            "service-user":  {"replicas": 3},
        }
        changes = compare_json_files(data, data, data, ['dev1', 'sit1'])
        assert changes == []

    # ── Scenario 2: Single scalar modify ────────────────────────
    def test_single_service_replicas_modified(self):
        """
        SCENARIO: service-admin replicas changed 2 → 4.
        WHAT IT TESTS: compare_json_files delegates to compare() which
        finds a scalar 'modify' at the "replicas" key.
        """
        le_old = {"service-admin": {"replicas": 2}}
        le_new = {"service-admin": {"replicas": 4}}
        he_old = {"service-admin": {"replicas": 2}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        assert len(changes) == 1
        assert changes[0][1] == 'modify'
        assert changes[0][0] == 'service-admin'

    # ── Scenario 3: Root object ADDED (new microservice) ────────
    def test_root_object_added_new_service(self):
        """
        SCENARIO: A brand-new service "service-payment" appears in le_new.
        WHAT IT TESTS: Lines 25-28 in compare_json_files — when a root
        key exists in le_new but NOT in le_old, it's a root object add.
        The entire service JSON is stored as `le_cur` and `he_cur` gets
        the env-transformed version.
        REAL-WORLD: A new microservice was onboarded to the dev env.
        """
        le_old = {"service-admin": {"replicas": 2}}
        le_new = {
            "service-admin": {"replicas": 2},
            "service-payment": {
                "replicas": 1,
                "image": {"image_name": "gcr.io/project/payment:1.0.0-b1-dev1"},
            },
        }
        he_old = {"service-admin": {"replicas": 2}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        added = [c for c in changes if c[7] == 'root object added']
        assert len(added) == 1
        assert added[0][0] == 'service-payment'
        assert added[0][1] == 'add'

    # ── Scenario 4: Root object DELETED (service decommissioned) ─
    def test_root_object_deleted_service_removed(self):
        """
        SCENARIO: service-auth was in le_old but is MISSING from le_new.
        WHAT IT TESTS: Lines 32-35 — root key in le_old but not le_new
        → records a 'delete' change with comment 'root object deleted'.
        REAL-WORLD: A microservice was decommissioned.
        """
        le_old = {
            "service-admin": {"replicas": 2},
            "service-auth": {"replicas": 1},
        }
        le_new = {"service-admin": {"replicas": 2}}
        he_old = {
            "service-admin": {"replicas": 2},
            "service-auth": {"replicas": 1},
        }
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        deleted = [c for c in changes if c[7] == 'root object deleted']
        assert len(deleted) == 1
        assert deleted[0][0] == 'service-auth'
        assert deleted[0][1] == 'delete'

    # ── Scenario 5: Multiple services with mixed changes ────────
    def test_multiple_services_mixed_changes(self):
        """
        SCENARIO: 3 services — one modified, one added, one deleted.
        WHAT IT TESTS: All branches of compare_json_files fire together.
        Validates correct attribution of changes to service names.
        """
        le_old = {
            "service-admin": {"replicas": 2},
            "service-auth": {"replicas": 1},
        }
        le_new = {
            "service-admin": {"replicas": 4},     # modified
            "service-new": {"replicas": 1},        # added
            # service-auth: deleted
        }
        he_old = {
            "service-admin": {"replicas": 2},
            "service-auth": {"replicas": 1},
        }
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        names = [c[0] for c in changes]
        assert 'service-admin' in names
        assert 'service-new' in names
        assert 'service-auth' in names

    # ── Scenario 6: Deeply nested key change ─────────────────────
    def test_nested_key_change(self):
        """
        SCENARIO: resources.limits.cpu changed "500m" → "1000m".
        WHAT IT TESTS: Recursive compare() traverses multiple levels.
        """
        le_old = {"svc": {"resources": {"limits": {"cpu": "500m"}}}}
        le_new = {"svc": {"resources": {"limits": {"cpu": "1000m"}}}}
        he_old = {"svc": {"resources": {"limits": {"cpu": "500m"}}}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        assert len(changes) == 1
        assert changes[0][1] == 'modify'

    # ── Scenario 7: Env var list — add entry ─────────────────────
    def test_env_var_added_to_list(self):
        """
        SCENARIO: A new env var is added to the "env" list of a service.
        WHAT IT TESTS: compare() → compare_list_of_dicts() path.
        Lists of dicts with "name" key use name-based matching.
        """
        le_old = {"svc": {"env": [{"name": "A", "value": "1"}]}}
        le_new = {"svc": {"env": [{"name": "A", "value": "1"}, {"name": "B", "value": "2"}]}}
        he_old = {"svc": {"env": [{"name": "A", "value": "1"}]}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        added = [c for c in changes if c[1] == 'add']
        assert len(added) == 1

    # ── Scenario 8: Env var list — delete entry ──────────────────
    def test_env_var_deleted_from_list(self):
        """
        SCENARIO: Env var "B" removed from the service's env list.
        WHAT IT TESTS: compare_list_of_dicts 'delete' branch.
        """
        le_old = {"svc": {"env": [{"name": "A", "value": "1"}, {"name": "B", "value": "2"}]}}
        le_new = {"svc": {"env": [{"name": "A", "value": "1"}]}}
        he_old = {"svc": {"env": [{"name": "A", "value": "1"}, {"name": "B", "value": "2"}]}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        deleted = [c for c in changes if c[1] == 'delete']
        assert len(deleted) == 1

    # ── Scenario 9: Env var list — modify entry ──────────────────
    def test_env_var_modified_in_list(self):
        """
        SCENARIO: Env var "A" value changed "1" → "999".
        WHAT IT TESTS: compare_list_of_dicts 'modify' branch.
        """
        le_old = {"svc": {"env": [{"name": "A", "value": "1"}]}}
        le_new = {"svc": {"env": [{"name": "A", "value": "999"}]}}
        he_old = {"svc": {"env": [{"name": "A", "value": "1"}]}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        modified = [c for c in changes if c[1] == 'modify']
        assert len(modified) == 1

    # ── Scenario 10: Pending addition (env var added earlier) ────
    def test_pending_addition_env_var(self):
        """
        SCENARIO: Env var "PENDING_VAR" exists in BOTH le_old and le_new,
        but NOT in he_old.
        WHAT IT TESTS: The key was added in a prior release but never
        promoted to the higher env → 'pending add' / 'Promotion Pending'.
        REAL-WORLD: Dev team added a config 2 releases ago, it was
        never promoted to SIT.
        """
        le_old = {"svc": {"env": [{"name": "PENDING_VAR", "value": "x"}]}}
        le_new = {"svc": {"env": [{"name": "PENDING_VAR", "value": "x"}]}}
        he_old = {"svc": {"env": []}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        pending = [c for c in changes if 'pending' in c[1].lower() or 'Pending' in c[7]]
        assert len(pending) >= 1

    # ── Scenario 11: Pending modify (value changed but never promoted) ─
    def test_pending_modify_scenario(self):
        """
        SCENARIO: A key "replicas" was changed from 2→4 in dev, and the
        higher env still has 2.  But le_old already had 4 (changed in
        a prior release).  The current le_new still has 4.
        WHAT IT TESTS: Since le_old == le_new but le_new differs from
        he_old, this falls into the 'pending add' (for dict keys) path.
        The compare engine sees the key exists in le_old and le_new
        but NOT in he_old with the same value → 'Promotion Pending'.
        """
        le_old = {"svc": {"new_key": "value"}}
        le_new = {"svc": {"new_key": "value"}}
        he_old = {"svc": {}}  # higher env doesn't have this key
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        pending = [c for c in changes if 'pending' in c[1].lower()]
        assert len(pending) >= 1

    # ── Scenario 12: Hotfix — image tag from dev to sit ──────────
    def test_hotfix_image_tag_dev_to_sit(self):
        """
        SCENARIO: Image tag changed due to a hotfix build.
        WHAT IT TESTS: The image change is detected as a 'modify'.
        The actual tag transformation is handled by update_image_tag()
        during write_changes_to_excel, but the comparison correctly
        identifies the image change.
        REAL-WORLD: Hotfix build 1.0.0-b15-dev1 pushed to dev.
        """
        le_old = {"svc": {"image": {"image_name": "gcr.io/proj/svc:1.0.0-b10-dev1"}}}
        le_new = {"svc": {"image": {"image_name": "gcr.io/proj/svc:1.0.0-b15-dev1"}}}
        he_old = {"svc": {"image": {"image_name": "gcr.io/proj/svc:1.0.0-b10-dev1"}}}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        modified = [c for c in changes if c[1] == 'modify']
        assert len(modified) == 1

    # ── Scenario 13: Empty configs (first-ever deployment) ───────
    def test_empty_le_old_all_new_services(self):
        """
        SCENARIO: le_old is completely empty (first deployment).
        WHAT IT TESTS: Every service in le_new is treated as a root-level
        addition.  This simulates the very first release to an env.
        """
        le_old = {}
        le_new = {
            "service-admin": {"replicas": 2},
            "service-user": {"replicas": 3},
        }
        he_old = {}
        changes = compare_json_files(le_old, le_new, he_old, ['dev1', 'sit1'])
        assert all(c[7] == 'root object added' for c in changes)
        assert len(changes) == 2


# ═══════════════════════════════════════════════════════════════════
# 2. update_image_tag() — IMAGE TAG TRANSFORMATION
# ═══════════════════════════════════════════════════════════════════

class TestUpdateImageTag:
    """
    update_image_tag(imagestr, lower_env, higher_env)

    LOGIC:
    - If lower_env is 'dev1' or 'dev2' → APPEND higher_env to tag
    - Otherwise → REPLACE lower_env with higher_env in tag
    - If no ':' in string → return as-is
    """

    def test_dev1_to_sit1_appends(self):
        """
        SCENARIO: Promoting from dev1 to sit1.
        WHAT IT TESTS: Lines 152-156.
        Tag "8.0.0-b2-release-r7-v2" → "8.0.0-b2-release-r7-v2-sit1"
        The higher env is APPENDED because dev is the first env.
        """
        result = update_image_tag('"gcr.io/proj/svc:1.0.0-b10"', 'dev1', 'sit1')
        assert 'sit1' in result

    def test_dev2_to_sit1_appends(self):
        """
        SCENARIO: Promoting from dev2 to sit1.
        WHAT IT TESTS: Same branch as dev1 (line 152 checks both).
        """
        result = update_image_tag('"gcr.io/proj/svc:2.0.0-b5"', 'dev2', 'sit1')
        assert 'sit1' in result

    def test_sit1_to_uat1_replaces(self):
        """
        SCENARIO: Promoting from sit1 to uat1.
        WHAT IT TESTS: Lines 158-163.
        Tag "1.0.0-b10-sit1" → "1.0.0-b10-uat1"
        The lower_env substring is REPLACED with higher_env.
        """
        result = update_image_tag('"gcr.io/proj/svc:1.0.0-b10-sit1"', 'sit1', 'uat1')
        assert 'uat1' in result
        assert 'sit1' not in result

    def test_uat1_to_prod_replaces(self):
        """
        SCENARIO: Promoting from uat1 to prod.
        WHAT IT TESTS: Same replacement logic for uat→prod.
        """
        result = update_image_tag('"gcr.io/proj/svc:1.0.0-b10-uat1"', 'uat1', 'prod')
        assert 'prod' in result
        assert 'uat1' not in result

    def test_no_colon_returns_as_is(self):
        """
        SCENARIO: Image string has no ':' (malformed).
        WHAT IT TESTS: Line 140-141 guard.
        """
        result = update_image_tag('gcr.io/proj/svc', 'dev1', 'sit1')
        assert result == 'gcr.io/proj/svc'

    def test_already_has_higher_env_no_duplicate(self):
        """
        SCENARIO: Tag already ends with 'sit1' when promoting dev1→sit1.
        WHAT IT TESTS: Line 154 — `if parts[-1] != higher_env` guard.
        Should NOT append 'sit1' again.
        """
        result = update_image_tag('"gcr.io/proj/svc:1.0.0-b10-sit1"', 'dev1', 'sit1')
        assert result.count('sit1') == 1

    def test_complex_tag_with_multiple_hyphens(self):
        """
        SCENARIO: Tag "8.0.0-b2-release-r7-v2" from dev1 to sit1.
        WHAT IT TESTS: Tag with many hyphen-separated parts.
        The higher env should be appended at the very end.
        """
        result = update_image_tag('"gcr.io/proj/svc:8.0.0-b2-release-r7-v2"', 'dev1', 'sit1')
        assert result.endswith('sit1')

    def test_hotfix_tag_transformation(self):
        """
        SCENARIO: Hotfix image "1.0.0-b15-hotfix-dev1" promoted sit→uat.
        WHAT IT TESTS: The lower_env 'sit1' is found in parts and replaced.
        REAL-WORLD: A critical bug fix was deployed as a hotfix.
        """
        result = update_image_tag('"gcr.io/proj/svc:1.0.0-b15-hotfix-sit1"', 'sit1', 'uat1')
        assert 'uat1' in result
        assert 'sit1' not in result


# ═══════════════════════════════════════════════════════════════════
# 3. update_image_repo_in_json_string() — JSON BLOB TRANSFORM
# ═══════════════════════════════════════════════════════════════════

class TestUpdateImageRepoInJsonString:
    """
    update_image_repo_in_json_string(json_str, lower, higher)

    Parses a JSON string, finds data["image"]["image_name"],
    and applies update_image_tag() to it.
    """

    def test_valid_json_with_image_name(self):
        """
        SCENARIO: JSON blob is a root object with an embedded image.
        WHAT IT TESTS: Lines 174-179 — finds and transforms image_name.
        REAL-WORLD: When a new service is added as a root object, the
        entire JSON blob needs its image tag transformed for the higher env.
        """
        data = {"image": {"image_name": "gcr.io/proj/svc:1.0.0-b1-dev1"}, "replicas": 2}
        json_str = json.dumps(data)
        result = update_image_repo_in_json_string(json_str, 'dev1', 'sit1')
        parsed = json.loads(result)
        assert 'sit1' in parsed['image']['image_name']

    def test_json_without_image_key(self):
        """
        SCENARIO: JSON doesn't have an "image" key.
        WHAT IT TESTS: Lines 175-176 guard — no 'image' key → return as-is.
        """
        data = {"replicas": 2, "resources": {"cpu": "500m"}}
        json_str = json.dumps(data)
        result = update_image_repo_in_json_string(json_str, 'dev1', 'sit1')
        assert result == json_str  # unchanged

    def test_invalid_json_returns_as_is(self):
        """
        SCENARIO: Input is not valid JSON (e.g., plain string).
        WHAT IT TESTS: Lines 170-172 — JSONDecodeError → return as-is.
        """
        result = update_image_repo_in_json_string("not-a-json", 'dev1', 'sit1')
        assert result == "not-a-json"

    def test_image_key_without_image_name(self):
        """
        SCENARIO: JSON has "image" key but no "image_name" sub-key.
        WHAT IT TESTS: Line 176 — `'image_name' in image_obj` check fails.
        """
        data = {"image": {"repository": "gcr.io/proj/svc"}, "replicas": 2}
        json_str = json.dumps(data)
        result = update_image_repo_in_json_string(json_str, 'dev1', 'sit1')
        assert result == json_str  # unchanged


# ═══════════════════════════════════════════════════════════════════
# 4. write_changes_to_excel() — EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════

class TestWriteChangesToExcel:
    """
    write_changes_to_excel(changes, release_note_path, envs)

    Writes the list of change tuples to an Excel file.
    Handles:
    - Normal-sized values → direct cell write
    - Values > 32,767 chars → saves to .txt file with hyperlink
    - image//image_name keys → calls update_image_tag()
    - root object added → calls update_image_repo_in_json_string()
    """

    def test_normal_changes_written(self, tmp_path):
        """
        SCENARIO: A few normal changes (under 32K chars each).
        WHAT IT TESTS: Lines 104-115 — standard append to worksheet.
        Verifies correct headers and data rows in the output Excel.
        """
        changes = [
            ('svc-admin', 'modify', 'replicas', '4', '2', '4', '2', 'Modified'),
            ('svc-user', 'add', 'cpu', '"500m"', '', '"500m"', '', 'Added'),
        ]
        envs = ['dev1', 'sit1']
        write_changes_to_excel(changes, str(tmp_path), envs)

        # Find the generated file
        xlsx_files = [f for f in os.listdir(tmp_path) if f.endswith('.xlsx')]
        assert len(xlsx_files) == 1

        wb = load_workbook(os.path.join(str(tmp_path), xlsx_files[0]))
        ws = wb.active
        assert ws.title == 'sit1'
        assert ws.cell(1, 1).value == 'Service name'
        assert ws.cell(2, 1).value == 'svc-admin'
        assert ws.cell(3, 1).value == 'svc-user'

    def test_no_changes_skips_file_creation(self, tmp_path):
        """
        SCENARIO: Empty changes list.
        WHAT IT TESTS: Lines 40-42 — early return, no file created.
        REAL-WORLD: When two releases are identical (no config changes).
        """
        write_changes_to_excel([], str(tmp_path), ['dev1', 'sit1'])
        xlsx_files = [f for f in os.listdir(tmp_path) if f.endswith('.xlsx')]
        assert len(xlsx_files) == 0

    def test_large_value_saved_to_txt(self, tmp_path):
        """
        SCENARIO: le_cur value exceeds 32,767 characters (Excel cell limit).
        WHAT IT TESTS: Lines 89-103 — large data saved to .txt file,
        hyperlink placed in Excel cell instead.
        REAL-WORLD: When a service has hundreds of env vars, the JSON
        string exceeds Excel's cell limit.
        """
        large_value = 'x' * 33000
        changes = [
            ('large-svc', 'add', '', large_value, '', large_value, '', 'root object added'),
        ]
        write_changes_to_excel(changes, str(tmp_path), ['dev1', 'sit1'])

        # Check .txt file was created
        txt_files = [f for f in os.listdir(tmp_path) if f.endswith('.txt')]
        assert len(txt_files) == 1
        assert txt_files[0] == 'large-svc.txt'

    def test_image_tag_transformation_on_write(self, tmp_path):
        """
        SCENARIO: A change with key='image//image_name' triggers tag transform.
        WHAT IT TESTS: Lines 80-82 — image tag transformation during write.
        The he_cur value should have the higher env appended/replaced.
        """
        changes = [
            ('svc', 'modify', 'image//image_name',
             '"gcr.io/proj/svc:1.0.0-b10-dev1"', '"gcr.io/proj/svc:1.0.0-b5-dev1"',
             '"gcr.io/proj/svc:1.0.0-b10-dev1"', '', 'Modified'),
        ]
        write_changes_to_excel(changes, str(tmp_path), ['dev1', 'sit1'])

        xlsx_files = [f for f in os.listdir(tmp_path) if f.endswith('.xlsx')]
        wb = load_workbook(os.path.join(str(tmp_path), xlsx_files[0]))
        ws = wb.active
        he_cur_cell = ws.cell(2, 6).value  # sit1-current value
        assert 'sit1' in he_cur_cell

    def test_root_object_added_transforms_image(self, tmp_path):
        """
        SCENARIO: Root object added with embedded image.image_name.
        WHAT IT TESTS: Lines 85-86 — update_image_repo_in_json_string()
        is called to transform image tags inside the full JSON blob.
        """
        data = {"image": {"image_name": "gcr.io/proj/svc:1.0.0-b1-dev1"}, "replicas": 1}
        json_str = json.dumps(data)
        changes = [
            ('new-svc', 'add', '', json_str, '', json_str, '', 'root object added'),
        ]
        write_changes_to_excel(changes, str(tmp_path), ['dev1', 'sit1'])

        xlsx_files = [f for f in os.listdir(tmp_path) if f.endswith('.xlsx')]
        wb = load_workbook(os.path.join(str(tmp_path), xlsx_files[0]))
        ws = wb.active
        he_cur_cell = ws.cell(2, 6).value
        assert 'sit1' in he_cur_cell

    def test_dict_value_serialized_safely(self, tmp_path):
        """
        SCENARIO: A change value is a raw dict (not a string).
        WHAT IT TESTS: Lines 45-50 — safe_excel_value() converts dicts
        to JSON strings before writing to Excel cells.
        """
        changes = [
            ('svc', 'modify', 'resources', {"cpu": "500m"}, '', {"cpu": "500m"}, '', 'Modified'),
        ]
        write_changes_to_excel(changes, str(tmp_path), ['dev1', 'sit1'])

        xlsx_files = [f for f in os.listdir(tmp_path) if f.endswith('.xlsx')]
        wb = load_workbook(os.path.join(str(tmp_path), xlsx_files[0]))
        ws = wb.active
        val = ws.cell(2, 4).value  # le_cur
        assert isinstance(val, str)
        assert 'cpu' in val


# ═══════════════════════════════════════════════════════════════════
# 5. touch_helmignore() — .HELMIGNORE GENERATION
# ═══════════════════════════════════════════════════════════════════

class TestTouchHelmignore:
    """
    touch_helmignore(meta_sheet_path, higher_env, helmignore_path)

    Reads metasheet from master branch to get environment names.
    Writes .helmignore excluding all envs EXCEPT the target higher env.
    """

    def test_read_metasheet_and_write_helmignore(self, sample_meta_sheet, tmp_path):
        """
        SCENARIO: Meta-sheet has columns [dev1, dev2, sit1, uat1, prod].
        Promoting to sit1 → .helmignore should exclude dev1, dev2, uat1, prod.
        WHAT IT TESTS: Lines 122-131.
        The function reads row 1 of the meta-sheet, iterates env values,
        and writes lines for every env != higher_env.
        REAL-WORLD: This is called during create_release_note to ensure
        Helm only packages the target env's values.
        """
        helmignore = tmp_path / ".helmignore"
        touch_helmignore(str(sample_meta_sheet), 'sit1', str(helmignore))

        content = helmignore.read_text()
        assert 'dev1-values/' in content
        assert 'dev2-values/' in content
        assert 'uat1-values/' in content
        assert 'prod-values/' in content
        assert 'sit1-values/' not in content

    def test_helmignore_for_prod(self, sample_meta_sheet, tmp_path):
        """
        SCENARIO: Promoting to prod → all other envs excluded.
        """
        helmignore = tmp_path / ".helmignore"
        touch_helmignore(str(sample_meta_sheet), 'prod', str(helmignore))

        content = helmignore.read_text()
        assert 'prod-values/' not in content
        assert 'dev1-values/' in content


# ═══════════════════════════════════════════════════════════════════
# 6. create_release_note_dir() — DIRECTORY MANAGEMENT
# ═══════════════════════════════════════════════════════════════════

class TestCreateReleaseNoteDir:
    """
    create_release_note_dir(release_note_path)

    - If directory exists: delete all .xlsx files inside
    - If directory doesn't exist: create it
    """

    def test_creates_dir_if_missing(self, tmp_path):
        """
        SCENARIO: The release_note folder doesn't exist yet.
        WHAT IT TESTS: Lines 197-199 — os.makedirs() is called.
        """
        new_dir = tmp_path / "release_note"
        create_release_note_dir(str(new_dir))
        assert new_dir.exists()

    def test_deletes_old_xlsx_if_dir_exists(self, tmp_path):
        """
        SCENARIO: Folder exists with old release-note-*.xlsx files.
        WHAT IT TESTS: Lines 185-196 — iterates and deletes .xlsx files.
        Other files (e.g. .txt) should be left untouched.
        REAL-WORLD: Prevents stale release notes from leaking into
        the next release cycle.
        """
        d = tmp_path / "release_note"
        d.mkdir()
        (d / "release-note-old.xlsx").write_bytes(b'old')
        (d / "service-data.txt").write_text("keep me")
        create_release_note_dir(str(d))
        assert not (d / "release-note-old.xlsx").exists()
        assert (d / "service-data.txt").exists()

    def test_empty_dir_no_error(self, tmp_path):
        """
        SCENARIO: Folder exists but is empty.
        WHAT IT TESTS: The loop body simply doesn't execute.
        """
        d = tmp_path / "release_note"
        d.mkdir()
        create_release_note_dir(str(d))
        assert d.exists()
