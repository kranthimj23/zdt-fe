"""
Shared fixtures for all test tiers.
Creates in-memory / tmp_path-based fixtures for Excel, YAML, and JSON test data.
"""
import os
import json
import pytest
import yaml
from openpyxl import Workbook


# ──────────────────────────────────────────────────────────────────────
# YAML Fixtures
# ──────────────────────────────────────────────────────────────────────

@pytest.fixture
def sample_yaml_dir(tmp_path):
    """Create a directory with sample YAML service files."""
    d = tmp_path / "app-values"
    d.mkdir()

    # service-admin.yaml
    (d / "service-admin.yaml").write_text(yaml.dump({
        "app": {"name": "admin-app"},
        "image": {"image_name": "gcr.io/project/admin:1.0.0-b10-dev1"},
        "replicas": 2,
        "resources": {"cpu": "500m", "memory": "512Mi"},
    }))

    # service-user.yaml
    (d / "service-user.yaml").write_text(yaml.dump({
        "app": {"name": "user-app"},
        "image": {"image_name": "gcr.io/project/user:2.0.0-b20-dev1"},
        "replicas": 3,
    }))

    # service-auth.yaml
    (d / "service-auth.yaml").write_text(yaml.dump({
        "app": {"name": "auth-app"},
        "image": {"image_name": "gcr.io/project/auth:3.0.0-b30-dev1"},
        "replicas": 1,
    }))

    return d


@pytest.fixture
def sample_yaml_dir_modified(tmp_path):
    """Same services as sample_yaml_dir but with modifications, additions, and deletions."""
    d = tmp_path / "app-values-modified"
    d.mkdir()

    # service-admin.yaml — MODIFIED (replicas changed, cpu changed)
    (d / "service-admin.yaml").write_text(yaml.dump({
        "app": {"name": "admin-app"},
        "image": {"image_name": "gcr.io/project/admin:1.1.0-b15-dev1"},
        "replicas": 4,
        "resources": {"cpu": "1000m", "memory": "512Mi"},
    }))

    # service-user.yaml — UNCHANGED
    (d / "service-user.yaml").write_text(yaml.dump({
        "app": {"name": "user-app"},
        "image": {"image_name": "gcr.io/project/user:2.0.0-b20-dev1"},
        "replicas": 3,
    }))

    # service-auth.yaml — DELETED (not present)
    # service-new.yaml — ADDED
    (d / "service-new.yaml").write_text(yaml.dump({
        "app": {"name": "new-app"},
        "image": {"image_name": "gcr.io/project/new:1.0.0-b1-dev1"},
        "replicas": 1,
    }))

    return d


@pytest.fixture
def large_yaml_content():
    """Generate a YAML content string that exceeds 32,767 characters (Excel cell limit)."""
    base = {"app": {"name": "large-service"}, "image": {"image_name": "gcr.io/project/large:1.0.0-b1-dev1"}}
    # Add enough env vars to push it over 32K
    env_list = []
    for i in range(500):
        env_list.append({"name": f"ENV_VAR_{i:04d}", "value": f"value_{i:04d}_{'x' * 50}"})
    base["env"] = env_list
    return base


# ──────────────────────────────────────────────────────────────────────
# Excel Fixtures
# ──────────────────────────────────────────────────────────────────────

@pytest.fixture
def sample_meta_sheet(tmp_path):
    """Create a meta-sheet.xlsx with environment columns and branch data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row: environments
    ws.append(["dev1", "dev2", "sit1", "uat1", "prod"])

    # Data rows: branch names (X = not promoted yet)
    ws.append(["release/1.0.0", "X", "X", "X", "X"])
    ws.append(["release/2.0.0", "release/2.0.0", "X", "X", "X"])

    path = tmp_path / "meta-sheet.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_release_note_excel(tmp_path):
    """Create a release-note Excel with add/modify/delete/pending rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "sit1"

    # Header row matching create_release_note.py format
    ws.append([
        'Service name', 'Change Request', 'Key',
        'dev1-current value', 'dev1-previous value',
        'sit1-current value', 'sit1-previous value',
        'Comment'
    ])

    # Row 2: MODIFY a scalar key
    ws.append([
        'service-admin', 'modify', 'replicas',
        '4', '2', '4', '2', 'Modified'
    ])

    # Row 3: ADD a new key
    ws.append([
        'service-admin', 'add', 'resources//cpu',
        '"1000m"', '', '"1000m"', '', 'Added'
    ])

    # Row 4: DELETE a key
    ws.append([
        'service-old', 'delete', 'replicas',
        '', '3', '', '3', 'Deleted'
    ])

    # Row 5: Root object added (new service)
    ws.append([
        'service-new', 'add', '',
        json.dumps({"app": {"name": "new-app"}, "replicas": 1}), '',
        json.dumps({"app": {"name": "new-app"}, "replicas": 1}), '',
        'root object added'
    ])

    # Row 6: Root object deleted
    ws.append([
        'service-auth', 'delete', '',
        '', json.dumps({"app": {"name": "auth-app"}}),
        '', json.dumps({"app": {"name": "auth-app"}}),
        'root object deleted'
    ])

    # Row 7: PENDING add (env var)
    ws.append([
        'service-user', 'pending', 'env//env',
        json.dumps({"name": "NEW_VAR", "value": "abc"}), '',
        json.dumps({"name": "NEW_VAR", "value": "abc"}), '',
        'Promotion Pending'
    ])

    # Row 8: MODIFY image tag
    ws.append([
        'service-admin', 'modify', 'image//image_name',
        '"gcr.io/project/admin:1.1.0-b15-dev1"', '"gcr.io/project/admin:1.0.0-b10-dev1"',
        '"gcr.io/project/admin:1.1.0-b15-sit1"', '',
        'Modified'
    ])

    path = tmp_path / "release-note-verified.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_deployment_yaml(tmp_path):
    """Create a minimal Helm deployment.yaml template for testing."""
    d = tmp_path / "helm-charts" / "templates"
    d.mkdir(parents=True)
    content = """\
apiVersion: apps/v1
kind: Deployment
spec:
  template:
    spec:
      containers:
        - name: {{ .Values.app.name }}
          image: {{ .Values.image.image_name }}
          env:
          {{- with .Values.env.service_admin }}
            {{- toYaml . | nindent 12 }}
          {{- end }}
          {{- with .Values.env.service_user }}
            {{- toYaml . | nindent 12 }}
          {{- end }}
      ports:
        - containerPort: 8080
"""
    (d / "deployment.yaml").write_text(content)
    return tmp_path


@pytest.fixture
def sample_infra_excel(tmp_path):
    """Create an infra_sheet.xlsx with multiple sheets for drift testing."""
    wb = Workbook()

    # Sheet: bucket
    ws_bucket = wb.active
    ws_bucket.title = "bucket"
    ws_bucket.append(["object_id", "bucket_name", "storage_class", "object_id_1"])
    ws_bucket.append([1, "my-bucket-1", "STANDARD", 1])
    ws_bucket.append([2, "my-bucket-2", "NEARLINE", 2])

    # Sheet: pull_subscription
    ws_ps = wb.create_sheet("pull_subscription")
    ws_ps.append(["object_id", "name", "topic", "ack_deadline_seconds", "object_id_1"])
    ws_ps.append([1, "sub-1", "topic-1", 30, 1])

    path = tmp_path / "infra_sheet.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_tmpl_dir(tmp_path):
    """Create template files for generate_tfvars."""
    d = tmp_path / "tmpl"
    d.mkdir()

    (d / "global.tmpl").write_text(
        'project_id   = "<<PROJECT_ID>>"\nregion       = "<<region>>"\n'
    )
    (d / "bucket.tmpl").write_text(
        'buckets = {\n<<bucket_name>>  = {\n        storage_class = "<<storage_class>>"\n  },\n}\n'
    )
    (d / "pull_subscription.tmpl").write_text(
        'pull_subscriptions = {\n<<pull_subscriptions>> = {\n    name = "<<name>>"\n    topic = "<<topic>>"\n  },\n}\n'
    )

    return d
