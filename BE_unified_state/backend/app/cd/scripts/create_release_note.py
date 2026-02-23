import os
import sys
import subprocess
import json
import openpyxl
from openpyxl import Workbook, load_workbook
import datetime

from utilities.dir_helpers import make_dir
from utilities.helpers import tokenize_url, get_parent_path, create_upgrade_services_txt
# from utilities.excel_helpers import write_changes_to_excel
from utilities.discarded_features import create_release_note_summary
from utilities.json_and_yaml_helpers import dump_and_replace, prepare_data, fetch_json, copy_missing_yaml_files
from utilities.git_helpers import clone_single_branch_and_checkout, stage_commit_and_push
from utilities.garuda_engine import compare, compare_list_of_dicts

envs = []
changes = []


def compare_json_files(le_old_data, le_new_data, he_old_data):
    changes = []
 
    # Check for changes in the JSON files
    for root in le_new_data.keys():
        if root not in le_old_data:
            modified_json = dump_and_replace(le_new_data[root], lower_env=envs[0], higher_env=envs[1])
            changes.append((root, 'add', '', json.dumps(le_new_data[root], indent=4), '', modified_json, '', 'root object added'))
        else:
            compare(le_old_data[root], le_new_data[root], root, changes, he_old_data[root])
 
    for root in le_old_data.keys():
        if root not in le_new_data:
            modified_json = dump_and_replace(le_old_data[root],lower_env=envs[0], higher_env=envs[1])
            changes.append((root, 'delete', '', '',json.dumps(le_old_data[root], indent=4), modified_json,'', 'root object deleted'))

    return changes

def write_changes_to_excel(changes, release_note_path, envs):
    if not changes:
        print("No differences found; skipping the creation of release note.")
        return

    # Helper to safely convert dicts to JSON strings for Excel
    def safe_excel_value(val):
        if isinstance(val, dict):
            return json.dumps(val)
        if val is None:
            return ''
        return val

    # Constant part of the filename
    base_filename = "release-note"
    # Get the current date and time
    now = datetime.datetime.now()
    # Format the date and time as a string
    date_time_str = now.strftime("%d-%b-%Y-%H-%M-%S")
    # Create the complete filename
    excel_file = f"{base_filename}-{date_time_str}.xlsx"
    excel_file_path = os.path.join(release_note_path, excel_file)

    # Check if the file exists
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
    else:
        wb = Workbook()

    ws = wb.active
    ws.title = envs[1]  # Name the first sheet

    ws.append([
        'Service name', 'Change Request', 'Key', f'{envs[0]}-current value', f'{envs[0]}-previous value', f'{envs[1]}-current value', f'{envs[1]}-previous value','Comment'])

    # Write the changes for the environment
    for change in changes:
        # print(change)
        service_name, change_type, key, le_cur, le_prev, he_cur, he_prev, comment = change

        # 🔄 If key is image//repository, update he_cur with transformed tag
        if key == 'image//image_name':
            he_cur = update_image_tag(he_cur, envs[0], envs[1]) # replace after 2nd hyphen
            he_cur = f'"{he_cur}"'

        # 🧠 Scenario 2: Root object with embedded image.repository
        elif isinstance(comment, str) and comment.strip().lower() == 'root object added':
            he_cur = update_image_repo_in_json_string(he_cur, envs[0], envs[1])

        # Check if value exceeds 32,767 characters
        if isinstance(le_cur, str) and len(le_cur) > 32767:
            # Save large data to a text file
            #solve for character exceed limit - shantanu tk
            he_cur = update_image_repo_in_json_string(he_cur, envs[0], envs[1])

            txt_file_name = f"{service_name}.txt"
            txt_file_path = os.path.join(release_note_path, txt_file_name)
            with open(txt_file_path, 'w') as txt_file:
                txt_file.write(he_cur)

            # Format path for hyperlink (absolute path)
            txt_file_path_linked = f'file:///{os.path.abspath(txt_file_path)}'.replace(' ', '%20')

            # Create a hyperlink in Excel pointing to this text file
            ws.append([service_name, change_type, key, txt_file_path_linked, '', txt_file_path_linked, '', comment]) #f'=HYPERLINK("{txt_file_path_linked}")'
        else:
            # Append new changes directly to the relevant sheet
            ws.append([
                service_name,
                change_type,
                key,
                safe_excel_value(le_cur),
                safe_excel_value(le_prev),
                safe_excel_value(he_cur),
                safe_excel_value(he_prev),
                comment
            ])

    # Save the updated workbook
    wb.save(excel_file_path)


def touch_helmignore(meta_sheet_file_path, higher_env, helmignore_path):
    workbook = openpyxl.load_workbook(meta_sheet_file_path)
    sheet = workbook.active
    env_values = [cell.value for cell in sheet[1]]

    try:
        with open(helmignore_path, 'w') as file:
            for env_value in env_values:
                if env_value != higher_env:
                    line = f"{env_value}-values/\n"
                    file.write(line)
    except FileNotFoundError:
        print(f"File not found: {helmignore_path}")
    except Exception as e:
        print(f"Error processing file: {e}")


def update_image_tag(imagestr, lower_env, higher_env):
    # split the image string into image name and tag parts
    if ':' not in imagestr:
        return imagestr
    
    image_name, tag = imagestr.split(':', 1)
    
    # Regex to find environment parts, assuming env tags are separated by dashes
    # For example: "8.0.0-b2-release-r7-v2"
    image_name = image_name.strip('"')
    parts = tag.split('-')
    parts[-1] = parts[-1].strip('"')

    # If lower environment is 'dev', we append higher_env (e.g., 'sit') at the end
    if lower_env == 'dev1' or lower_env == 'dev2':
        # Add higher_env to end if not already present
        if parts[-1] != higher_env:
            parts.append(higher_env)
        updated_tag = '-'.join(parts)
    else:
        # If lower_env is other than dev, replace the lower_env with higher_env
        # Find the index of lower_env in parts and replace it
        if lower_env in parts:
            idx = parts.index(lower_env)
            parts[idx] = higher_env
        updated_tag = '-'.join(parts)
    
    return f'{image_name}:{updated_tag}'


def update_image_repo_in_json_string(json_str, lower_env, higher_env):
    try:
        data = json.loads(json_str)
    except json.JSONDecodeError:
        return json_str  # return as-is if not a valid JSON

    if isinstance(data, dict):
        image_obj = data.get('image', {})
        if isinstance(image_obj, dict) and 'image_name' in image_obj:
            image_obj['image_name'] = update_image_tag(image_obj['image_name'], lower_env, higher_env)
            data['image'] = image_obj
            return json.dumps(data, indent=4)

    return json_str


def create_release_note_dir(release_note_path):
    if os.path.isdir(release_note_path):
        print(f"Checking folder: {release_note_path}")
        # List all files in the directory
        for filename in os.listdir(release_note_path):
            # Check if the file ends with .xlsx
            if filename.endswith('.xlsx'):
                file_path = os.path.join(release_note_path, filename)
                try:
                    os.remove(file_path)  # Delete the file
                    print(f"Deleted: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {e}")
    else:
        print(f"Folder does not exist: {release_note_path}, hence creating one")
        os.makedirs(release_note_path)


def main():
    repos_info = {
        'mb-helmcharts': sys.argv[5]
    }
 
    for repo_name, repo_url in repos_info.items():
        promote_branch_x_1 = sys.argv[1]
        promote_branch_x = sys.argv[2]
        target_folder_x_1 = make_dir("generate-config/promotion-x-1", f"{repo_name}")
        target_folder_x = make_dir("generate-config/promotion-x", f"{repo_name}")
    
    try:
        repo_url = tokenize_url(repo_url)
    except Exception as e:
        print(f"Logging error: {e}")
        raise

    envs.append(sys.argv[3].strip())
    envs.append(sys.argv[4].strip())
    
    target_folder = make_dir()
    meta_sheet_file_path = os.path.join(target_folder, f"meta-sheet.xlsx")
    clone_single_branch_and_checkout(repo_url, promote_branch_x_1, target_folder_x_1)
    clone_single_branch_and_checkout(repo_url, promote_branch_x, target_folder_x)
    clone_single_branch_and_checkout(repo_url, "master", target_folder)

    helmignore_path = os.path.join(target_folder_x, "helm-charts", ".helmignore")
    touch_helmignore(meta_sheet_file_path, envs[1], helmignore_path)

    release_note_path = os.path.join(target_folder_x, "helm-charts", f"{envs[1]}-values", "app-values", "release_note")
    create_release_note_dir(release_note_path)
 
    higher_env_x_1 = os.path.join(target_folder_x_1, f"helm-charts/{envs[1]}-values/app-values")
    lower_env_x = os.path.join(target_folder_x, f"helm-charts/{envs[0]}-values/app-values")
 
    if os.path.exists(higher_env_x_1) and os.path.exists(lower_env_x):
        '''
        Compares YAML files in two folders and copies missing files
        from the lower environment to the higher environment.
        '''
        copy_missing_yaml_files(higher_env_x_1, lower_env_x, envs[0], envs[1])

    le_x_1_json_data = prepare_data(target_folder_x_1, envs[0])
    le_x_json_data = prepare_data(target_folder_x, envs[0])
    he_x_1_json_data = prepare_data(target_folder_x_1, envs[1])

    changes = compare_json_files(le_x_1_json_data, le_x_json_data, he_x_1_json_data)
    write_changes_to_excel(changes, release_note_path, envs)

    yaml_path = os.path.dirname(fetch_json(target_folder_x, envs[0]))
    create_release_note_summary(yaml_path, target_folder_x, release_note_path, envs[1])
    
    create_upgrade_services_txt(release_note_path, envs[1], target_folder_x, envs[0])

    try:
        commit_message = f'Pushing the release_note into the branch: {sys.argv[2]}'
        stage_commit_and_push(repo_url, target_folder_x, sys.argv[2], commit_message, rebase=True)
        print("Push successful.")
    except subprocess.CalledProcessError as e:
        print("Git command failed!")
        print("Return code:", e.returncode)
        print("Command:", e.cmd)
        print("Output:", e.output)
        print("Error:", e.stderr)
        sys.exit(1)
 
    target_folder_x = os.path.dirname(target_folder_x)
    target_folder_x_1 = os.path.dirname(target_folder_x_1)
 
if __name__ == '__main__':
    main()