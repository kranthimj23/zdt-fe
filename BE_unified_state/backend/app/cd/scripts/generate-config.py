import os
import openpyxl
from openpyxl import load_workbook
import subprocess
import sys

from utilities.dir_helpers import make_dir
from utilities.json_and_yaml_helpers import read_yaml_files_to_json
from utilities.excel_helpers import get_sheet, get_cell_value, get_headers, get_sheets_with_values
from utilities.helpers import tokenize_url, extract_hyperlink_path
from utilities.git_helpers import clone_repo_and_checkout, stage_commit_and_push
from utilities.json_and_yaml_helpers import save_json_to_file, create_yaml_files_from_json, try_parse_json
from utilities.deployment_helpers import update_txt_file_with_yaml_values, insert_hardcoded_value, modify_deployment_yaml
from utilities.garuda_engine import handle_data_env
 
update_template= []
deleted_services = []


def get_release_note(target_folder_x, higher_env):
    release_note_path = os.path.join(target_folder_x, "helm-charts", f"{higher_env}-values", "app-values", f"release_note")

    foldernames = os.listdir(target_folder_x)
 
    for foldername in foldernames:
        if foldername == "helm-charts":
            for filename in os.listdir(release_note_path):
                if filename.startswith("release-note") and "verified" in filename:
                    excel_file_path = os.path.join(release_note_path,filename)
                    print("Excel file path: ", excel_file_path)
    return excel_file_path


def apply_changes_to_json(json_data, excel_file_path, sheet_name, lower_env, higher_env):
    wb = load_workbook(excel_file_path, data_only=True)
    ws = wb[sheet_name]
 
    # Read header row to map column names to indices
    #6: sit1-current-values
    headers = {}
    headers = get_headers(headers, ws)

    # Iterate over data rows starting from row 2
    for row_num in range(2, ws.max_row + 1):
        service_name = get_cell_value(row_num, 'Service name', headers, ws)
        change_request = get_cell_value(row_num, 'Change Request', headers, ws)
        key = get_cell_value(row_num, 'Key', headers, ws)
        comments = get_cell_value(row_num, 'Comment', headers, ws)
        le_cur = get_cell_value(row_num, f'{lower_env}-current value', headers, ws)
        le_prev = get_cell_value(row_num, f'{lower_env}-previous value', headers, ws)
        he_cur = get_cell_value(row_num, f'{higher_env}-current value', headers, ws)
        he_prev = get_cell_value(row_num, f'{higher_env}-previous value', headers, ws)
 
        # Handle addition and deletion of a root object
        if key is None or (isinstance(key, str) and key.strip() == ""):
            print(service_name, comments, "root user added/deleted")

            comments = comments.strip().lower()
            if comments == "root object deleted":
                print("root object to be deleted found")
                if service_name and service_name in json_data:
                    deleted_services.append(service_name)
                    del json_data[service_name]
                    print(f"Deleted root object: {service_name}")
                else:
                    print("Did not enter the loop", service_name)
                continue
            elif comments == "root object added":
                new_root_object = service_name
                if new_root_object not in json_data:
                    json_data[new_root_object] = {}
                    print(f"Added new root object: {new_root_object}")
                if he_cur is not None and (not (isinstance(he_cur, str) and he_cur.strip() == "")):
                    if isinstance(he_cur, str) and he_cur.startswith('file'):
                        hyperlink_path = extract_hyperlink_path(he_cur)
                        hyperlink_path = rf'/{hyperlink_path}'
                        with open(hyperlink_path, 'r') as txt_file:
                            he_cur = txt_file.read().strip() 
                    json_data[new_root_object] = try_parse_json(he_cur)
                update_template.append(service_name)
                continue
            elif service_name in ['data', 'env']:
                pass  # Skip the key checks for 'data' and 'env'
            else:
                print(f"Error: Key is missing or empty in row {row_num}.")
                raise ValueError(f"Missing or empty key encountered in row {row_num}.")
 
        if he_cur is None:
            parsed_value = try_parse_json(he_prev)
        else:
            parsed_value = try_parse_json(he_cur)

        if parsed_value is None:
            print(f"Skipping row {row_num} due to None parsed_value")
            continue  # Skip processing this row if parsed_value is None

        if service_name in ['data', 'env']:
            handle_data_env(json_data, service_name, change_request, parsed_value)
            continue

        # General handling for other services with keys
        key_path = key.split('//')
        obj = json_data.setdefault(service_name, {})

        for k in key_path[:-1]:
            if not isinstance(obj, dict):
                obj = {}
            obj = obj.setdefault(k, {})

        final_key = key_path[-1]

        if key_path[0] == "env" or (len(key_path) > 1 and key_path[1] == "env"):
            if change_request == 'add':
                if final_key not in obj:
                    obj[final_key] = parsed_value
                obj[final_key].insert(-1, parsed_value)
                print(f"Added to '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'modify':
                for entry in obj[final_key]:
                    if entry['name'] == parsed_value['name']:
                        entry['value'] = parsed_value['value']
                        print(f"Modified '{final_key}' entry in '{service_name}': {parsed_value}")
                        break
            elif change_request == 'delete':
                if final_key in obj and parsed_value is not None:
                    if isinstance(parsed_value, dict):
                        names_to_delete = {parsed_value[k] for k in parsed_value.keys() if k == 'name'}
                    else:
                        names_to_delete = {item['name'] for item in parsed_value}
                    # names_to_delete = {item['name'] for item in parsed_value}
                    print(f"Deleted entry from '{final_key}' in '{service_name}'.")
                    obj[final_key] = [ 
                        entry for entry in obj[final_key] if not (isinstance(entry, dict) and 'name' in entry and entry['name'] in names_to_delete)
                    ]
                    if len(obj[final_key]) == 0:
                        del obj[final_key]
            elif change_request == 'pending':
                if final_key not in obj:
                    obj[final_key] = parsed_value
                else:
                    obj[final_key].insert(-1, parsed_value)  # Same logic as 'add'
                print(f"PENDING→ADDED to '{final_key}' in '{service_name}': {parsed_value}")

        else:
            if change_request == 'modify':
                obj[final_key] = parsed_value
                print(f"Modified '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'pending':
                obj[final_key] = parsed_value
                print(f"PENDING→ADDED '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'add':
                obj[final_key] = parsed_value
                print(f"Added '{final_key}' in '{service_name}': {parsed_value}")
            elif change_request == 'delete' and final_key in obj:
                current_value = obj[final_key]
                if isinstance(current_value, list) and isinstance(parsed_value, dict) and 'name' in parsed_value:
                    obj[final_key] = [
                        item for item in current_value if not (isinstance(item, dict) and item.get('name') == parsed_value['name'])
                    ]
                    print(f"Deleted matching entry from list '{final_key}' in '{service_name}'.")
                else:
                    del obj[final_key]
                    print(f"Deleted '{final_key}' from '{service_name}'.")

        # Check for missing or empty value
        if he_cur is None or (isinstance(he_cur, str) and he_cur.strip() == ""):
            if change_request == "delete":
                continue
            else:
                raise ValueError(f"Missing or empty value encountered in row {row_num}.")

    return json_data


def apply_sed_to_yaml(folder_path):
    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        return
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".yaml"):
            file_path = os.path.join(folder_path, file_name)
            sed_command = f"sed -i -E -e 's/([[:space:]]+value:[[:space:]]*)(0[0-9]+)([[:space:]]*$)/\\1\"\\2\"\\3/' -e 's/([[:space:]]+value:[[:space:]]*)([A-Z])([[:space:]]*$)/\\1\"\\2\"\\3/' {file_path}"
 
            try:
                subprocess.run(sed_command, shell=True, check=True)
                # print(f"Processed: {file_name}")
            except subprocess.CalledProcessError as e:
                print(f"Error processing {file_name}: {e}")

 
def create_txt_file(excel_path, env, txt_path, output_folder):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[env]  # Use sheet argument correctly
    max_row = ws.max_row
    values_to_be_added = set()
 
    for row in range(2, max_row + 1):  # Start from the second row
        name = ws.cell(row=row, column=1).value
        second_col_value = ws.cell(row=row, column=2).value
 
        if name is None:
            continue
 
        elif name in ('data', 'env'):
            continue
 
        elif second_col_value == "delete":
            # Skip writing this name to file
            continue
        else:
            values_to_be_added.add(name)
 
    with open(txt_path, 'w') as file:
        for i in sorted(values_to_be_added):
            file.write(f"{i}\n")
    
    update_txt_file_with_yaml_values(txt_path, output_folder)


def update_meta_sheet(lower_env, higher_env, promote_branch, repo_url):
    master_dir = make_dir(dir_name='master')
 
    try:
        clone_repo_and_checkout(repo_url, "master", master_dir)
        # Process Excel file
        meta_sheet_file_path = os.path.join(master_dir, f"meta-sheet.xlsx")
        ws = get_sheet(meta_sheet_file_path)
        # Find column indices
        headers = [cell.value for cell in ws[1]]
        lower_col = headers.index(lower_env) + 1
        higher_col = headers.index(higher_env) + 1
 
        # Update logic
        updated = False
        for row in ws.iter_rows(min_row=2):  # Skip header
            if row[lower_col - 1].value == promote_branch:
                if row[higher_col - 1].value == promote_branch:
                    print("already updated")
                else:
                    row[higher_col - 1].value = promote_branch
                updated = True
 
                break
        if not updated:
            raise ValueError(f"{promote_branch} not found in {lower_env} column")
 
        workbook = ws.parent
        workbook.save(meta_sheet_file_path)
        print("saved excel file")
 
        # Commit and push changes to master branch
        stage_commit_and_push(repo_url, master_dir, 'master', f'Promoted {promote_branch} from {lower_env} to {higher_env}')
 
    except subprocess.CalledProcessError as e:
        print(f"Git operation failed: {e}")
        print(f"Return code: {e.returncode}")
        print(f"Command: {e.cmd}")
        print(f"Output: {e.output}")
        print(f"Stderr: {e.stderr}")
        return False


def main():
    repos_info = {
        'mb-helmcharts': sys.argv[5]
    }

    #look into this loop for whats happening exactly
    for repo_name, repo_url in repos_info.items():
        promote_branch_x_1 = sys.argv[1]
        promote_branch_x = sys.argv[2]
        target_folder_x_1 = make_dir("generate-config/promotion-x-1", f"{repo_name}")
        target_folder_x = make_dir("generate-config/promotion-x", f"{repo_name}")

    try:
        repo_url = tokenize_url(repo_url)
    except Exception as e:
        print(f"Logging error: {e}")
        raise

    clone_repo_and_checkout(repo_url, promote_branch_x_1, target_folder_x_1)
    clone_repo_and_checkout(repo_url, promote_branch_x, target_folder_x)

    
    lower_env = sys.argv[3]
    higher_env = sys.argv[4]

    release_note_file_path = get_release_note(target_folder_x, higher_env)
    sheet = get_sheets_with_values(release_note_file_path, higher_env)

    app_values_path = f"helm-charts/{sheet}-values/app-values" 
    folder_path = os.path.join(target_folder_x_1, app_values_path)
    initial_output_file = os.path.join(target_folder_x_1, app_values_path, f"config-{sheet}.json")
    updated_output_file = os.path.join(target_folder_x, app_values_path, f"config-{sheet}.json")
    output_folder = os.path.join(target_folder_x, app_values_path)
    txt_file_path = os.path.join(target_folder_x, app_values_path, f"{sheet}.txt" )

    json_data = read_yaml_files_to_json(folder_path)
    save_json_to_file(json_data, initial_output_file)

    updated_json = apply_changes_to_json(json_data, release_note_file_path, sheet, lower_env, higher_env)

    insert_hardcoded_value(target_folder_x, update_template, higher_env)
    modify_deployment_yaml(target_folder_x, deleted_services, higher_env)
    save_json_to_file(updated_json, updated_output_file)
    create_yaml_files_from_json(updated_output_file, output_folder)
    apply_sed_to_yaml(output_folder)
    create_txt_file(release_note_file_path, sheet, txt_file_path, output_folder)
            
    try:
        commit_message = f'Config files generated: {sys.argv[2]}'
        stage_commit_and_push(repo_url, target_folder_x, sys.argv[2], commit_message, rebase=True)
        print("Push successful.")
        update_meta_sheet(lower_env, sheet, promote_branch_x, repo_url)
    except subprocess.CalledProcessError as e:
        print("Git command failed!")
        print("Return code:", e.returncode)
        print("Command:", e.cmd)
        print("Output:", e.output)
        print("Error:", e.stderr)
        sys.exit(1)

    target_folder_x = os.path.dirname(target_folder_x)
    target_folder_x_1 = os.path.dirname(target_folder_x_1)
 
if __name__ == "__main__":
    main()