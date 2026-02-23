import os
import subprocess
import shutil
from openpyxl import load_workbook, Workbook
import tempfile
import re
import sys

from git_helpers import (
    inject_git_token,
    clone_single_branch_and_checkout,
    clone_repo_and_checkout,
    is_base_branch_exists,
    configure_git_user,
    stage_commit_and_push,
    stage_specific_files_commit_and_push,
)
 
def add_value_to_columns(file_path, sheet_name, file_name, branch_name):
    if os.path.exists(file_path):
    # Load the workbook
        workbook = load_workbook(file_path)
        print("AQL__SQL",file_path)
    else:
        workbook = Workbook()
        # Remove the default sheet created by openpyxl
        default_sheet = workbook.active
        workbook.remove(default_sheet)
 
    # Check if the sheet exists
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the workbook.")
        sheet = workbook.create_sheet(sheet_name)
        print(sheet)
        # Add headers to the first row
        sheet.cell(row=1, column=1, value="Filename")
        sheet.cell(row=1, column=2, value="Branchname")
    else:
        sheet = workbook[sheet_name]

 
 
    # Find the column indices for "Filename" and "Branch"
    filename_col_index = None
    branch_col_index = None
 
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == "Filename":
            filename_col_index = col[0].column  # Get column index for "Filename"
        elif col[0].value == "Branchname":
            branch_col_index = col[0].column  # Get column index for "Branch"
 
    # Check if both columns exist
    if filename_col_index is None:
        print(f"Column 'Filename' not found in sheet '{sheet_name}'.")
    if branch_col_index is None:
        print(f"Column 'Branchname' not found in sheet '{sheet_name}'.")
 
    # Find the next empty row in the columns
    next_row = sheet.max_row + 1
 
    # Add values to respective columns
    sheet.cell(row=next_row, column=filename_col_index).value = file_name
    sheet.cell(row=next_row, column=branch_col_index).value = branch_name
    print(sheet.cell(row=next_row, column=filename_col_index).value)
    print(sheet.cell(row=next_row, column=branch_col_index).value)
 
    # Save the workbook
    workbook.save(file_path)
    print(f"Values added to columns {file_name}' and {branch_name}' in sheet '{sheet_name}'.")
 
 
def update_meta_sheet(lower_env, higher_env, promote_branch, repo_url):
    repo_path = tempfile.mkdtemp()

    print(os.path.dirname(repo_path))
    try:
        repo_url = inject_git_token(repo_url)
        clone_single_branch_and_checkout(repo_url, 'master', repo_path)

        if not is_base_branch_exists(repo_url, promote_branch):
            raise ValueError(f"Branch {promote_branch} does not exist in repository")

        print("Step1")
        print(os.path.dirname(repo_path))
        excel_path = os.path.join(repo_path, 'meta-sheet.xlsx')
        wb = load_workbook(excel_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        lower_col = headers.index(lower_env) + 1
        higher_col = headers.index(higher_env) + 1

        updated = False
        for row in ws.iter_rows(min_row=2):
            if row[lower_col-1].value == promote_branch:
                row[higher_col-1].value = promote_branch
                updated = True
                break
        print("Step2")
        if not updated:
            raise ValueError(f"{promote_branch} not found in {lower_env} column")

        wb.save(excel_path)

        configure_git_user(repo_path)
        stage_specific_files_commit_and_push(
            repo_path,
            'master',
            f'Promote {promote_branch} from {lower_env} to {higher_env}',
            ['meta-sheet.xlsx'],
            pull_before_push=False,
        )

        shutil.rmtree(repo_path, ignore_errors=True)

        return True

    except subprocess.CalledProcessError as e:
        print(f"Git operation failed: {e.stderr}")
        return False

    except Exception as e:
        print(f"An error occurred: {e}")
        return False
 
 
def get_branch_range(lower_env, higher_env):
    temp_dir = tempfile.mkdtemp()

    clone_single_branch_and_checkout(github_url, 'master', temp_dir)
    for i in os.listdir(temp_dir):
        if i.startswith("meta-sheet"):
            file_path = os.path.join(temp_dir,i)
 
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    sheet = wb.active
 
    # Get column indices for lower_env and higher_env
    headers = [cell.value for cell in sheet[1]]
    lower_env_col = headers.index(lower_env) + 1
    higher_env_col = headers.index(higher_env) + 1
 
    # Find upper limit (last non-X value in lower_env column)
    lower_env_values = [
        sheet.cell(row=i, column=lower_env_col).value
        for i in range(2, sheet.max_row + 1)
        if sheet.cell(row=i, column=lower_env_col).value != 'X'
    ]
    upper_limit = lower_env_values[-1]
 
    # Find lower limit
    higher_env_values = [
        sheet.cell(row=i, column=higher_env_col).value
        for i in range(2, sheet.max_row + 1)
        if sheet.cell(row=i, column=higher_env_col).value != 'X'
    ]
    last_higher_env_branch = higher_env_values[-1]
 
    try:
        idx = lower_env_values.index(last_higher_env_branch)
        lower_limit = lower_env_values[idx + 1]
    except (ValueError, IndexError):
        raise Exception("Could not determine lower limit")
 
    print(lower_limit)
 
    # Create the ordered list
    branch_list = []
    collect = False
    for value in lower_env_values:
        if value == lower_limit:
            collect = True
        if collect:
            branch_list.append(value)
        if value == upper_limit:
            break
 
    for i in branch_list:
        if i.endswith("-Err"):
            branch_list.remove(i)
    shutil.rmtree(temp_dir, ignore_errors=True)
    return branch_list
 
 
def merge_branches(branches, promotion_branch, aql_folder, sql_folder, lower_env, higher_env):
    try:
        # Fetch all branches from all remotes
        print("Fetching all branches from all remotes...")
        subprocess.run(["git", "fetch", "--all"])
 
        try:
            subprocess.run(["git", "fetch", "origin", promotion_branch])
        except subprocess.CalledProcessError as e:
            print(f"Error fetching the {promotion_branch}: {e}")
            return  # Exit if fetching the branch fails
 
        # Check if promotion branch exists locally
        try:
            subprocess.run(["git", "checkout", promotion_branch])
            #subprocess.run(["git", "switch", promotion_branch])x
            print(f"Successfully switched to {promotion_branch}")
        except subprocess.CalledProcessError as e:
            print(f"Error checking out {promotion_branch}: {e}")
            return
 
        # Create AQL folder if it doesn't exist
        if not os.path.exists(aql_folder):
            os.makedirs(aql_folder)
            print(aql_folder)
 
        # Create SQL folder if it doesn't exist
        if not os.path.exists(sql_folder):
            os.makedirs(sql_folder)
 
        # Function to copy files of a specific type
        def merge_files(file_type, folder, branches, dest_branch):
            """
            Copies multiple files of a specific type from the specified folders in each branch
            into multiple files in the corresponding folder of the destination branch.
 
            Parameters:
            - file_type: Type of files to copy (e.g., ".aql", ".sql")
            - folder: Path to the folder containing files to copy (e.g., aql_folder, sql_folder)
            - branches: List of branch names to fetch files from
            - dest_branch: Destination branch where files will be copied
            """
            # Initialize a dictionary to hold file contents
            file_contents = {}
 
            if len(branches) > 1:
                count = 1
                for branch in branches:
                    try:
                        # Checkout the branch to access its files
                        subprocess.run(["git", "checkout", branch])
                        print(f"Successfully switched to {branch}")
 
                        # Get a list of files of the specified type in the folder
                        files = [f for f in os.listdir(folder) if f.endswith(file_type)]
 
                        for file in files:
                            # Construct the full path to the file
                            file_path = os.path.join(folder, file)
 
                            # Read the file content
                            with open(file_path, 'r') as f:
                                content = f.read()
 
                            # Store the content with the branch and file name
                            new_file_name = f"{count}_{file}"
                            file_contents[new_file_name] = content
                            if file_type.endswith(".aql"):
                                sheetname = "AQL"
                            elif file_type.endswith(".sql"):
                                sheetname = "SQL"
                            add_value_to_columns(release_note, sheetname, new_file_name, branch)
 
                            print(f"Read content from {file} in {branch}")
 
                    except subprocess.CalledProcessError as e:
                        print(f"Error checking out {branch}: {e}")
 
                    count = count+1
 
            else:
 
                try:
                    # Checkout the branch to access its files
                    subprocess.run(["git", "checkout", branches[0]])
                    print(f"Successfully switched to {branches[0]}")
 
                    # Get a list of files of the specified type in the folder
                    files = [f for f in os.listdir(folder) if f.endswith(file_type)]
 
                    print(files)
 
                    for file in files:
                        # Construct the full path to the file
                        file_path = os.path.join(folder, file)
 
                        # Read the file content
                        with open(file_path, 'r') as f:
                            content = f.read()
 
                        # Store the content with the branch and file name
                        new_file_name = f"{file}"
                        file_contents[new_file_name] = content
                        if file_type.endswith(".aql"):
                            sheetname = "AQL"
                        elif file_type.endswith(".sql"):
                            sheetname = "SQL"
                        add_value_to_columns(release_note, sheetname, new_file_name, branches[0])
 
                        print(f"Read content from {file} in {branches[0]}")
 
                except subprocess.CalledProcessError as e:
                    print(f"Error checking out {branches[0]}: {e}")
 
            print(file_contents)
 
            # Switch back to the destination branch
            subprocess.run(["git", "checkout", dest_branch])
            print(f"Successfully switched back to {dest_branch}")
            folder = folder.replace(lower_env, higher_env)
            txt_file = os.path.join(folder, f'{higher_env}.txt')
 
            # Flag to track if any content was written to txt_file
            file_written = False
 
            # Write the contents to new files
            for file_name, content in file_contents.items():
                modified_content = content.replace(lower_env, higher_env)
 
                # If the content is actually modified (or non-empty), write to file
                if modified_content:
                    file_path = os.path.join(folder, file_name)

                    # Ensure the directory exists
                    os.makedirs(os.path.dirname(file_path), exist_ok=True)

                    with open(file_path, 'w') as f:
                        f.write(modified_content)
                    print(f"Wrote content to {file_path}")
 
                    # Append the file_name to txt_file
                    with open(txt_file, 'a') as f:
                        f.write(file_name + '\n')
                    file_written = True  # Mark that we wrote to txt_file
 
            # Define the sorting function
            def sort_file_names(file_path):
                # Read file names from the text file
                with open(file_path, 'r') as file:
                    lines = file.readlines()
 
                # Strip newline characters
                lines = [line.strip() for line in lines]
 
                # Custom sorting function
                def custom_sort_key(s):
                    match = re.match(r'^(\d+)_', s)
                    if match:
                        num = int(match.group(1))
                        return (0, num, s)  # Sort by number then alphabetically
                    else:
                        return (1, s)  # Push non-numeric to end, sort alphabetically
 
                # Sort the lines
                sorted_lines = sorted(lines, key=custom_sort_key)
 
                # Write the sorted lines back to the file
                with open(file_path, 'w') as file:
                    for line in sorted_lines:
                        file.write(line + '\n')
 
            # Call sort_file_names only if txt_file exists and was written to
            if file_written and os.path.exists(txt_file):
                sort_file_names(txt_file)
            else:
                print(f"No files modified, so {txt_file} was not created or updated.")
 
            subprocess.run(["git", "add", folder])
            print(f"Added all new files in {folder} to staging")

        merge_files(".aql", aql_folder, branches, promotion_branch)

        configure_git_user(os.getcwd())
        stage_commit_and_push(
            os.getcwd(),
            promotion_branch,
            'Merged AQL and SQL files from multiple branches',
            pull_before_push=False,
        )

        print("Merge and push complete!")

    except subprocess.CalledProcessError as e:
        print(f"Error: {e}")
        print(e.output)
 
 
 
if __name__ == "__main__":
    main_temp_dir = tempfile.mkdtemp()
    github_url = sys.argv[1]
    lower_env = sys.argv[2]
    higher_env = sys.argv[3]
    branches_to_merge = get_branch_range(lower_env, higher_env)
    promotion_branch_name = branches_to_merge[-1]
    aql_folder_name = rf'helm-charts/{lower_env}-values/db-scripts/AQL'
    sql_folder_name = rf'helm-charts/{lower_env}-values/db-scripts/SQL'
    release_note_path = sys.argv[4]
    release_note = os.path.join(main_temp_dir, release_note_path)
    print(release_note)
    github_url = inject_git_token(github_url)
    clone_single_branch_and_checkout(github_url, promotion_branch_name, main_temp_dir)
    os.chdir(main_temp_dir) #change directory to target
    merge_branches(branches_to_merge, promotion_branch_name, aql_folder_name, sql_folder_name,lower_env,higher_env)
    #update_meta_sheet(lower_env, higher_env, promotion_branch_name, github_url)
    print("Scripts executed")
