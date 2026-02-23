import os
import sys
import subprocess
import json
import yaml
import openpyxl
from openpyxl import Workbook, load_workbook
import shutil
import datetime
from openpyxl.utils import get_column_letter
import re

from utilities.git_helpers import clone_single_branch_and_checkout, stage_commit_and_push
from utilities.helpers import tokenize_url

 
envs = []
 
 
def copy_missing_yaml_files(higher_env_x_1, lower_env_x, lower_env, higher_env):
    """
    Compares YAML files in two folders and copies missing files
    from the lower environment to the higher environment.
 
    Args:
        higher_env_x_1 (str): Path to the higher environment folder.
        lower_env_x (str): Path to the lower environment folder.
    """
 
    higher_env_x_1_files = set(os.listdir(higher_env_x_1))
    lower_env_x_files = set(os.listdir(lower_env_x))
 
    # Identify YAML files present in lower_env_x but not in higher_env_x_1
    missing_files = [f for f in lower_env_x_files if f.endswith(('.yaml', '.yml')) and f not in higher_env_x_1_files]
 
    for filename in missing_files:
        source_path = os.path.join(lower_env_x, filename)
        destination_path = os.path.join(higher_env_x_1, filename)
 
        try:
            shutil.copy2(source_path, destination_path)  # copy2 preserves metadata
 
        except Exception as e:
            print(f"Error copying {filename}: {e}")
 
        try:
            with open(destination_path, 'r') as file:
                file_content = file.read()

            # ✅ Now pass file_content to safe_env_replace
            updated_content = safe_env_replace(file_content, lower_env, higher_env)

            print("[DEBUG] YAML after safe_env_replace:")
            print(updated_content)

            with open(destination_path, 'w') as file:
                file.write(updated_content) # Write the updated content back to the file [2]
 
            # print(f"Successfully replaced '{lower_env}' with '{higher_env}' in '{filename}'.")
 
        except FileNotFoundError:
            print(f"Error: The file '{filename}' was not found.")
        except Exception as e:
            print(f"An error occurred: {e}")
 
    if not os.path.exists(higher_env_x_1):
        print(f"Error: The folder {higher_env_x_1} does not exist.")
    if not os.path.exists(lower_env_x):
        print(f"Error: The folder {lower_env_x} does not exist.")


def safe_env_replace(content: str, lower_env: str, higher_env: str) -> str:
    """
    Replaces tag suffixes (e.g., :14-dev → :14-sit) without changing registry domains (e.g., .pkg.dev).
    """
    # Replace tag suffix: e.g., :14-dev → :14-sit
    content = re.sub(
        rf'(:[\w\-]+-){re.escape(lower_env)}(\b)',
        rf'\1{higher_env}',
        content
    )

    # DO NOT replace 'dev' inside domain names like .pkg.dev
    # So we skip global replace

    return content       

 
def create_release_note_summary(files_path, target_folder_x, existing_release_note_dir,input_sheet_name):
    print("Executing create release note.py")
    # Find first Excel file in existing release note folder
    excel_file = None
    if os.path.exists(existing_release_note_dir):
        for f in os.listdir(existing_release_note_dir):
            if f.endswith(".xlsx"):
                excel_file = os.path.join(existing_release_note_dir, f)
                break
 
    if not excel_file:
        print("Excel sheet not found in the source release note directory.")
        return
 
    # Ensure new release note directory exists
    if not os.path.exists(target_folder_x):
        print("Path to summary sheet not specified")
 
    # Compose the directory in which the file is to be copied
    new_dir = os.path.join(target_folder_x, "helm-charts", f"{envs[1]}-values", "release_note")
 
    # Create directory if it doesn't exist
    os.makedirs(new_dir, exist_ok=True)
 
    # Then, compose the full new Excel file path
    new_excel_file = os.path.join(new_dir, "release-note-summary.xlsx")
    # Copy original Excel file to this new location
    try:
        shutil.copy2(excel_file, new_excel_file)
        print(f"Copied {excel_file} to {new_excel_file}")
    except Exception as e:
        print(f"Failed to copy Excel file: {e}")
        return
 
    # Continue with rest of your processing by loading new_excel_file...
    wb = openpyxl.load_workbook(new_excel_file)
 
    # Get YAML files from the specified path
    yaml_files = [f for f in os.listdir(files_path) if f.endswith('.yaml') or f.endswith('.yml')]
 
    # Check if "Summary" sheet already exists; if so, remove it
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
 
    # Create a new "Summary" sheet
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.sheet_state = "visible"
 
    # Add column headers including new "Build Number" at third position
    headers = ["Services", "Status", "Build Number", "Comments", "Owner"]
    summary_sheet.append(headers)
 
    # Add service names from YAML files
    for yaml_file in yaml_files:
        service_name = os.path.splitext(yaml_file)[0]  # Use file name (without extension) as service name
        # Initialize row with empty values for Status, Build Number, Comments, Owner
        summary_sheet.append([service_name, "", "", "", ""])
 
    # Add dropdown options to the "Status" column
    dropdown_options = ["Updated", "No modifications", "New service", "Deleted service"]
 
    # Data validation for dropdowns
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(dropdown_options)),
        allow_blank=True
    )
 
    # Add the dropdown to all cells in the "Status" column (from row 2 onwards)
    for row in range(2, summary_sheet.max_row + 1):
        cell = summary_sheet.cell(row=row, column=2)  # Status column
        dv.add(cell.coordinate)
 
    summary_sheet.add_data_validation(dv)
 
    # Load the specified sheet for comments
    if input_sheet_name in wb.sheetnames:
        comments_sheet = wb[input_sheet_name]
    else:
        print(f"Sheet '{input_sheet_name}' not found in the Excel file.")
        return
 
    # Function to extract build number from image_name string
    def extract_build_number(image_name_value):
        # Example: format x.y.z-bpqr-qwertyu-0.0
        # Goal: find substring starting with 'b' (like 'bpqr') and extract numeric portion (pqr)
        # This regex captures the numeric part following 'b' in the dash-separated substring after the last hyphen
        # but from your example, the number is after b in bpqr (where pqr is digits)
        # We'll extract the substring starting with b and then digits after it (with some leniency)
        # If digits can be anywhere after 'b', this regex helps extract first digits after b
 
        # A more concrete approach:
        # Find segment after first 'b' in the string where digits appear consecutively
 
        match = re.search(r"-b(\d+)", image_name_value) # '-' followed by 'b' and digits
        if match:
            return match.group(1)  # return digits after 'b'
        else:
            # Alternatively, more relaxed pattern: find 'b' followed by digits anywhere
            match = re.search(r"b(\d+)", image_name_value)
            if match:
                return match.group(1)
        return ""  # Return empty string if no match
 
    # Create a dictionary for comments to speed up lookup: {service_name: comment}
    comments_dict = {}
    for r in range(2, comments_sheet.max_row + 1):
        svc_name = comments_sheet.cell(row=r, column=1).value
        comment = comments_sheet.cell(row=r, column=comments_sheet.max_column).value  # Assuming last column has comment
        if svc_name:
            comments_dict[svc_name] = comment
 
    # Iterate through rows in Summary sheet to update Status and Build Number
    for row in range(2, summary_sheet.max_row + 1):
        service_name = summary_sheet.cell(row=row, column=1).value
 
        comment = comments_dict.get(service_name, None)
        # Update status based on comment
        status = "No modifications"
        if comment == "root object added":
            status = "New service"
        elif comment == "root object deleted":
            status = "Deleted service"
        elif comment in ["Modified", "Added"]:
            status = "Updated"
 
        summary_sheet.cell(row=row, column=2).value = status
 
        # Build Number column is at column 3 (index 3)
        build_number_cell = summary_sheet.cell(row=row, column=3)
 
        # If status is Updated or New service, parse YAML and extract build number
        if status in ["Updated", "New service"]:
            # Compose full path to yaml file
            yaml_file_path = os.path.join(files_path, f"{service_name}.yaml")
            if not os.path.exists(yaml_file_path):
                # Try .yml extension if .yaml does not exist
                yaml_file_path = os.path.join(files_path, f"{service_name}.yml")
 
            if os.path.exists(yaml_file_path):
                try:
                    with open(yaml_file_path, 'r') as yf:
                        yaml_content = yaml.safe_load(yf)
 
                    # Search for 'image_name' key anywhere in yaml (top level or nested)
                    # If your structure is nested, you may need a recursive search
                    def find_image_name(data):
                        if isinstance(data, dict):
                            for k, v in data.items():
                                if k == "image_name":
                                    return v
                                else:
                                    found = find_image_name(v)
                                    if found:
                                        return found
                        elif isinstance(data, list):
                            for item in data:
                                found = find_image_name(item)
                                if found:
                                    return found
                        return None
 
                    image_name_val = find_image_name(yaml_content)
                    if image_name_val:
                        build_num = extract_build_number(str(image_name_val))
                        build_number_cell.value = build_num
                    else:
                        build_number_cell.value = ""
 
                except Exception as e:
                    print(f"Failed to read/parse {yaml_file_path}: {e}")
                    build_number_cell.value = ""
            else:
                build_number_cell.value = ""
        else:
            build_number_cell.value = ""
 
    # Adjust column widths for better readability (columns A to E now)
    for col in range(1, 6):  # A=1 to E=5
        column_letter = get_column_letter(col)
        summary_sheet.column_dimensions[column_letter].width = 20
 
    # Save the workbook with the updated Summary sheet
    wb.save(new_excel_file)
 
    # Re-open and reorder sheets to put Summary first
    wb = openpyxl.load_workbook(new_excel_file)
 
    if "Summary" in wb.sheetnames:
        sheet = wb["Summary"]
 
        # Get the current index of the sheet
        current_index = wb._sheets.index(sheet)
 
        # If the sheet is not already the first one, move it
        if current_index != 0:
            # Remove the sheet from its current position
            wb._sheets.pop(current_index)
            # Insert the sheet at the beginning
            wb._sheets.insert(0, sheet)
 
        wb.save(new_excel_file)
 
 
 
def fetch_json(target_folder, env):
    json_path = ''
    foldername = f"{env}-values"
    for root, folders, files in os.walk(target_folder):
        if foldername in folders:
            lower_env_path = os.path.join(root, foldername, 'app-values')
            for filename in os.listdir(lower_env_path):
                if filename == f"config-{env}.json":
                    json_path = os.path.join(lower_env_path, filename)
                    break
            if not json_path:
                # Create an empty config-dev.json if it doesn't exist
                json_path = os.path.join(lower_env_path, f"config-{env}.json")
                with open(json_path, 'w') as f:
                    json.dump({}, f)  # Create an empty JSON object
            break  # Exit after processing the 'dev-values' folder
 
    return json_path
 
def yaml_to_json(folder_path):
    json_data = {}
    # Convert YAML files to JSON structure
    for filename in os.listdir(folder_path):
        if filename.endswith('.yaml'):
            yaml_path = os.path.join(folder_path, filename)
            with open(yaml_path, 'r') as yaml_file:
                yaml_content = yaml.safe_load(yaml_file)
                json_data[filename[:-5]] = yaml_content  # Use filename without .yaml as the root key
 
    return json_data
 
def get_parent_path(path: str) -> str:
    if not path:
        return path
    return path.rsplit("//", 1)[0] if "//" in path else path
 
def dump_and_replace(json_obj, lower_env, higher_env):
    json_str = json.dumps(json_obj, indent=4)
    return json_str
 

def compare_json_files(le_old_data, le_new_data, he_old_data):
    changes = []
 
    # Check for changes in the JSON files
    for root in le_new_data.keys():
        if root not in le_old_data:
            modified_json = dump_and_replace(le_new_data[root], lower_env=envs[0], higher_env=envs[1])
            changes.append((root, 'add', '', json.dumps(le_new_data[root], indent=4), '', modified_json, '', 'root object added'))
        else:
            compare(le_old_data[root], le_new_data[root], root, changes, he_old_data[root])
 
    for root in le_old_data.keys():
        if root not in le_new_data:
            modified_json = dump_and_replace(le_old_data[root],lower_env=envs[0], higher_env=envs[1])
            changes.append((root, 'delete', '', '',json.dumps(le_old_data[root], indent=4), modified_json,'', 'root object deleted'))
 
    return changes
 
def compare(le_old, le_new, root, changes, he_old, path=''):
    # Handle dictionary structures
    if isinstance(le_old, dict) and isinstance(le_new, dict):

        he_old = he_old or {}

        all_keys = set(he_old.keys()) | set(le_old.keys()) | set(le_new.keys())

        for k in all_keys:
            new_key_path = f"{path}//{k}" if path else k

            he_val = he_old.get(k)
            le_old_val = le_old.get(k)
            le_new_val = le_new.get(k)

            # 🔴 Deleted (exists in old lower but removed in new)
            if k in le_old and k not in le_new:
                modified_json = dump_and_replace(
                    le_old_val,
                    lower_env=envs[0],
                    higher_env=envs[1]
                )

                changes.append((
                    root, 'delete', get_parent_path(new_key_path),
                    '',
                    json.dumps(le_old_val, indent=4),
                    '',
                    json.dumps(he_val, indent=4) if he_val else '',
                    'Deleted'
                ))
            # 🔴 Deletion Pending (exists in higher, removed in new)
            elif k in he_old and k not in le_new:
                changes.append((
                    root, 'pending delete', get_parent_path(new_key_path),
                    '', '', '',
                    json.dumps(he_val, indent=4),
                    'Deletion Pending'
                ))

            # 🟢 Added / Promotion Pending
            elif k in le_new and k not in he_old:
                modified_json = dump_and_replace(
                    le_new_val,
                    lower_env=envs[0],
                    higher_env=envs[1]
                )

                if k not in le_old:
                    changes.append((
                        root, 'add', get_parent_path(new_key_path),
                        json.dumps(le_new_val, indent=4),
                        '',
                        modified_json,
                        '',
                        'Added'
                    ))
                else:
                    changes.append((
                        root, 'pending add', get_parent_path(new_key_path),
                        json.dumps(le_new_val, indent=4),
                        '',
                        modified_json,
                        '',
                        'Promotion Pending'
                    ))

            # 🔁 Recursive compare
            elif k in le_old and k in le_new:
                compare(
                    le_old_val,
                    le_new_val,
                    root,
                    changes,
                    he_val,
                    new_key_path
                )

    # Handle lists
    elif isinstance(le_old, list) and isinstance(le_new, list):
        if all(isinstance(i, dict) for i in le_old) and all(isinstance(i, dict) for i in le_new):
            # Check if both lists are not empty
            if le_old and le_new and "name" in le_old[0] and "name" in le_new[0]:
                compare_list_of_dicts(le_old, le_new, root, changes, he_old, path)
            else:
                # Handle lists without "name" key or if lists are empty
                for i, le_old_item in enumerate(le_old):
                    if i < len(le_new):
                        if le_old_item != le_new[i]:
                            modified_json = dump_and_replace(le_new[i], lower_env=envs[0], higher_env=envs[1])
                            changes.append((root, 'modify', f"{path}", "["+json.dumps(le_new[i], indent=4)+"]", "["+json.dumps(le_old_item, indent=4)+"]",  "["+modified_json+"]","["+json.dumps(he_old[i], indent=4)+"]",'Modified'))
                    else:
                        modified_json = dump_and_replace(le_old_item, lower_env=envs[0], higher_env=envs[1])
                        changes.append((root, 'delete', f"{path}", '',json.dumps(le_old_item, indent=4), '',json.dumps(he_old[i], indent=4),'Deleted'))
 
 
                # Add any new elements from the new list
                for i in range(len(le_old), len(le_new)):
                    modified_json = dump_and_replace(le_new[i], lower_env=envs[0], higher_env=envs[1])
                    changes.append((root, 'add', f"{path}", json.dumps(le_new[i], indent=4), '',modified_json,'', 'Added'))
 
    # Compare scalar values
    else:
        if le_old != le_new:
            path = get_parent_path(path)
            modified_json = dump_and_replace(le_new, lower_env=envs[0], higher_env=envs[1])
            changes.append((root, 'modify', path, json.dumps(le_new, indent=4), json.dumps(le_old, indent=4) ,modified_json,json.dumps(he_old, indent=4) ,'Modified'))
 
def compare_list_of_dicts(le_old_list, le_new_list, root, changes, he_old_list, path=''):

    le_old_dict = {i["name"]: i for i in le_old_list if "name" in i}
    le_new_dict = {i["name"]: i for i in le_new_list if "name" in i}
    he_old_dict = {i["name"]: i for i in he_old_list if "name" in i}

    all_keys = set(le_old_dict) | set(le_new_dict) | set(he_old_dict)

    for key in all_keys:
        le_old_item = le_old_dict.get(key)
        le_new_item = le_new_dict.get(key)
        he_old_item = he_old_dict.get(key)

        key_path = f"{path}//{key}"

        # 🔴 Deleted
        if key in le_old_dict and key not in le_new_dict:
            modified_json = dump_and_replace(
                le_old_item,
                lower_env=envs[0],
                higher_env=envs[1]
            )
            changes.append((
                root, 'delete', get_parent_path(key_path),
                '',
                json.dumps(le_old_item, indent=4),
                '',
                json.dumps(he_old_item, indent=4) if he_old_item else '',
                'Deleted'
            ))
        # 🔴 Deletion Pending
        elif key in he_old_dict and key not in le_new_dict:
            changes.append((
                root, 'pending delete', get_parent_path(key_path),
                '', '', '',
                json.dumps(he_old_item, indent=4),
                'Deletion Pending'
            ))

        # 🟢 Added / Promotion Pending
        elif key in le_new_dict and key not in he_old_dict:
            modified_json = dump_and_replace(
                le_new_item,
                lower_env=envs[0],
                higher_env=envs[1]
            )

            if key not in le_old_dict:
                changes.append((
                    root, 'add', get_parent_path(key_path),
                    json.dumps(le_new_item, indent=4),
                    '',
                    modified_json,
                    '',
                    'Added'
                ))
            else:
                changes.append((
                    root, 'pending add', get_parent_path(key_path),
                    json.dumps(le_new_item, indent=4),
                    '',
                    modified_json,
                    '',
                    'Promotion Pending'
                ))

        # 🔁 Modified
        elif le_old_item != le_new_item:
            modified_json = dump_and_replace(
                le_new_item,
                lower_env=envs[0],
                higher_env=envs[1]
            )
            changes.append((
                root, 'modify', get_parent_path(key_path),
                json.dumps(le_new_item, indent=4),
                json.dumps(le_old_item, indent=4),
                modified_json,
                json.dumps(he_old_item, indent=4) if he_old_item else '',
                'Modified'
            ))


def update_image_tag(image_str, env_keyword):
    """
    Safely update only the environment part of the image tag (e.g., :14-dev → :14-sit)
    without affecting the registry domain (e.g., .pkg.dev).
    """
    if not isinstance(image_str, str) or ":" not in image_str:
        return image_str

    image_repo, tag = image_str.rsplit(":", 1)
    tag_parts = tag.split("-")

    if len(tag_parts) < 2:
        return image_str  # no env suffix found

    # Replace last part (env) with new one
    tag_parts[-1] = env_keyword
    new_tag = "-".join(tag_parts)

    return f"{image_repo}:{new_tag}"


def update_image_repo_in_json_string(json_str_or_obj, env_keyword):
    """
    Safely update imageName, image_name, and tag fields in a JSON string or object,
    replacing environment parts of tags without affecting domains etc.
    Accepts either JSON string or already loaded Python object.
    Returns pretty-printed JSON string.
    """
    # If input is already dict or list, use it directly
    if isinstance(json_str_or_obj, (dict, list)):
        data = json_str_or_obj
    else:
        # Otherwise, try to parse JSON string
        try:
            data = json.loads(json_str_or_obj)
        except json.JSONDecodeError:
            # If invalid JSON string, return as-is
            return json_str_or_obj

    # Proceed only if data is a dict (since you're updating image field)
    if not isinstance(data, dict):
        return json.dumps(data, indent=4) if isinstance(data, (dict, list)) else data

    image_obj = data.get('image')
    if isinstance(image_obj, dict):
        # Update relevant keys safely
        for key in ['imageName', 'image_name', 'tag']:
            if key in image_obj and isinstance(image_obj[key], str):
                image_obj[key] = update_image_tag(image_obj[key], env_keyword)

        data['image'] = image_obj

    # Return the updated JSON as pretty string
    return json.dumps(data, indent=4)


def write_changes_to_excel(changes, release_note_path, envs, env_list):
    if not changes:
        print("No differences found; skipping the creation of release note.")
        return

    # Helper to safely convert dicts to JSON strings for Excel
    def safe_excel_value(val):
        if isinstance(val, (dict, list)):
            return json.dumps(val)
        if val is None:
            return ''
        return val

    # Constant part of the filename
    base_filename = "release-note"
    # Get the current date and time
    now = datetime.datetime.now()
    # Format the date and time as a string
    date_time_str = now.strftime("%d-%b-%Y-%H-%M-%S")
    # Create the complete filename
    excel_file = f"{base_filename}-{date_time_str}.xlsx"
    excel_file_path = os.path.join(release_note_path, excel_file)

    # Check if the file exists
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
    else:
        wb = Workbook()

    ws = wb.active
    ws.title = envs[1]  # Name the first sheet

    ws.append([
        'Service name', 'Change Request', 'Key', f'{envs[0]}-current value', f'{envs[0]}-previous value', f'{envs[1]}-current value', f'{envs[1]}-previous value','Comment'])

    # Write the changes for the environment
    for change in changes:
        print(change)
        service_name, change_type, key, le_cur, le_prev, he_cur, he_prev, comment = change

        # 🔄 If key is image//repository, update he_cur with transformed tag
        if key == 'image//image_name':
            he_cur = update_image_tag(he_cur, envs[1])  # replace after 2nd hyphen

        # 🧠 Scenario 2: Root object with embedded image.repository
        elif isinstance(comment, str) and comment.strip().lower() == 'root object added':
            he_cur = update_image_repo_in_json_string(he_cur, envs[1])

        # Check if value exceeds 32,767 characters
        if isinstance(le_cur, str) and len(le_cur) > 32767:
            # Save large data to a text file
            txt_file_name = f"{service_name}.txt"
            txt_file_path = os.path.join(release_note_path, txt_file_name)
            with open(txt_file_path, 'w') as txt_file:
                txt_file.write(le_cur)

            # Format path for hyperlink (absolute path)
            txt_file_path_linked = f'file:///{os.path.abspath(txt_file_path)}'.replace(' ', '%20')

            # Create a hyperlink in Excel pointing to this text file
            ws.append([service_name, change_type, key, f'=HYPERLINK("{txt_file_path_linked}")', '', '', '', comment])
        else:
            # Append new changes directly to the relevant sheet
            ws.append([
                service_name,
                change_type,
                key,
                safe_excel_value(le_cur),
                safe_excel_value(le_prev),
                safe_excel_value(he_cur),
                safe_excel_value(he_prev),
                comment
            ])

    # Save the updated workbook
    wb.save(excel_file_path)    

 
def get_input(prompt):
    attempts = 3
    for _ in range(attempts):
        value = input(prompt)
        if value:
            return value
        print("Input cannot be empty. Please try again.")
    print("Cannot proceed with execution as no input was entered.")
    exit(1)
 
 
def execute(target_folder_x, lower_env, higher_env, repo_url):
    try:
        release_note_path = os.path.join(target_folder_x, "helm-charts", f"{higher_env}-values", "release_note")
 
        print(release_note_path)
        if os.path.exists(release_note_path) and os.path.isdir(release_note_path):
            print(f"Checking folder: {release_note_path}")
 
            # List all files in the directory
            for filename in os.listdir(release_note_path):
                # Check if the file ends with .xlsx
                if filename.endswith('.xlsx'):
                    print(filename)
                    file_path = os.path.join(release_note_path, filename)
                    if os.path.exists(file_path):
                        print("This is the file path for db-scripts: ",file_path)
                else:
                    # Constant part of the filename
                    base_filename = "release-note"
                    # Get the current date and time
                    now = datetime.datetime.now()
                    # Format the date and time as a string
                    date_time_str = now.strftime("%d-%b-%Y-%H-%M-%S")
                    # Create the complete filename
                    excel_file = f"{base_filename}-{date_time_str}.xlsx"
                    file_path = os.path.join(release_note_path, excel_file)
                    print(file_path)
 
                pythonexec=os.getenv("PYTHON_EXEC", "python3.11")
                # result = subprocess.run(
                #     [pythonexec, sys.argv[6] ,repo_url, lower_env, higher_env,file_path],
                #     check=True,
                #     capture_output=True,
                #     text=True
                # )
                result = True
                return result
 
 
 
    except subprocess.CalledProcessError as e:
        print("Script failed to execute.")
        print("STDOUT:", e.stdout)
        print("STDERR:", e.stderr)
 
 
def create_upgrade_services_txt(excel_path, sheet_name, repo_root, lower_env):
    txt_path = os.path.join(repo_root, 'upgrade-services.txt')
 
    if os.path.exists(txt_path):
        os.remove(txt_path)
 
    # Identify the correct .xlsx file
    file_path = None
    for i in os.listdir(excel_path):
        if i.endswith(".xlsx"):
            file_path = os.path.join(excel_path, i)
            break
 
    if not file_path:
        print("No Excel file found in the given path.")
    else:
        wb = load_workbook(file_path)
 
        if sheet_name not in wb.sheetnames:
            print("Application release note does not exist")
        else:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            try:
                key_col = headers.index('Key') + 1
                service_name_col = headers.index('Service name') + 1
                value_col = headers.index(f'{lower_env}-current value') + 1
                comment_col = headers.index(f'Comment') + 1
            except ValueError as e:
                raise ValueError(f"Required column missing: {e}")
 
            with open(txt_path, 'w') as file:
                for row in ws.iter_rows(min_row=2):
                    key_cell = row[key_col - 1].value
                    service_name = row[service_name_col - 1].value
                    value = row[value_col - 1].value
                    comment = row[comment_col - 1].value
 
                    if key_cell and ('image//tag' in str(key_cell) or 'image//image_name' in str(key_cell)):
                        if service_name and value:
                            file.write(f"{service_name}:{value}\n")
                            print(f"{service_name}:{value}")
                    elif comment and comment.strip().lower() == "root object added" and service_name not in ("data", "env"):
                        if value:
                            try:
                                parsed = json.loads(value)
                                print(parsed)
                                tag = parsed.get("image", {}).get("image_name")
                                if service_name and tag:
                                    file.write(f"{service_name}:{tag}\n")
                                    print(f"{service_name}:{tag}")
                            except json.JSONDecodeError as e:
                                print(f"Failed to parse JSON for service {service_name}: {e}")
                        else:
                            print(f"Value is None for service {service_name} with 'root object added'")
                    else:
                        print(f"Skipping service {service_name}")
 
 
def main():
    repos_info = {
        'mb-helmcharts': sys.argv[5]
    }
 
    for repo_name, repo_url in repos_info.items():
        # promote_branch_x_1 = input("Enter the branch containing the stable release: ")
        promote_branch_x_1 = sys.argv[1]
        # promote_branch_x = input("Enter the branch containing the updated files for release: ")
        promote_branch_x = sys.argv[2]
        target_folder_x_1 = os.path.join(os.getcwd(), "generate-configs", "promotion-x-1", f"{repo_name}")
        target_folder_x = os.path.join(os.getcwd(), "generate-configs", "promotion-x", f"{repo_name}")
 
    envs.append(sys.argv[3].strip())
    envs.append(sys.argv[4].strip())
    print("The list of envs is",envs)
    env_list = ['dev2', 'sit2', 'uat2', 'prod']
    repos = sys.argv[6:]
    # Environment list (you can modify based on your use case)
    # envs = ['dev', 'sit', 'uat', 'preprod', 'perf', 'mig/dm', 'sec', 'prod']
 
    repo_url = tokenize_url(repo_url)

    clone_single_branch_and_checkout(repo_url, promote_branch_x_1, target_folder_x_1)
    clone_single_branch_and_checkout(repo_url, promote_branch_x, target_folder_x)
 
    # if envs[0].startswith('dev'):
    print("dev2 is the lower-env")
    #to create the release note folder
    release_note_path = os.path.join(target_folder_x, "helm-charts", f"{envs[1]}-values", "app-values", "release_note")
    if os.path.exists(release_note_path) and os.path.isdir(release_note_path):
        print(f"Checking folder: {release_note_path}")
 
        # List all files in the directory
        for filename in os.listdir(release_note_path):
            # Check if the file ends with .xlsx
            if filename.endswith('.xlsx'):
                file_path = os.path.join(release_note_path, filename)
                try:
                    os.remove(file_path)  # Delete the file
                    print(f"Deleted: {file_path}")
                except Exception as e:
                    print(f"Error deleting file {file_path}: {e}")
    else:
        print(f"Folder does not exist: {release_note_path}, hence creating one")
    if not os.path.exists(release_note_path):
        os.makedirs(release_note_path)
 
    higher_env_x_1 = os.path.join(target_folder_x_1, f"helm-charts/{envs[1]}-values/app-values")
    lower_env_x = os.path.join(target_folder_x, f"helm-charts/{envs[0]}-values/app-values")
 
    if os.path.exists(higher_env_x_1) and os.path.exists(lower_env_x):
        copy_missing_yaml_files(higher_env_x_1, lower_env_x, envs[0], envs[1])
 
    #To fetch the path of config-dev.json from promotion-x-1 branch
    le_previous_json_path = fetch_json(target_folder_x_1, envs[0])
    print(le_previous_json_path)
 
    #To fetch the path of config-dev.json from promotion-x branch
    le_new_json_path = fetch_json(target_folder_x, envs[0])
    print(le_new_json_path)
 
    # Load previous JSON data
    try:
        with open(le_previous_json_path, 'r') as le_old_file:
            le_old_data = json.load(le_old_file)
    except (json.JSONDecodeError, FileNotFoundError):
        print("Previous JSON file is invalid or not found.")
        le_old_data = {}
    print(os.path.dirname(le_previous_json_path))
    lower_env_path = os.path.dirname(le_previous_json_path)
    le_old_data = yaml_to_json(lower_env_path)
 
    # Save the previous JSON data
    with open(le_previous_json_path, 'w') as previous_json_file:
        json.dump(le_old_data, previous_json_file, indent=4)
 
    lower_env_path = os.path.dirname(le_new_json_path)
    le_json_data = yaml_to_json(lower_env_path)
 
    # Save the new JSON data
    with open(le_new_json_path, 'w') as le_new_json_file:
        json.dump(le_json_data, le_new_json_file, indent=4)
 
 
    he_previous_json_path = fetch_json(target_folder_x_1, envs[1])
    print(he_previous_json_path)
 
    higher_env_path = os.path.dirname(he_previous_json_path)
    he_old_data = yaml_to_json(higher_env_path)
 
            # Save the previous JSON data
    with open(he_previous_json_path, 'w') as he_previous_json_path:
        json.dump(he_old_data, he_previous_json_path, indent=4)
 
 
    # Compare JSON files
    changes = compare_json_files(le_old_data, le_json_data, he_old_data)
        # Write changes to Excel with multiple sheets
    write_changes_to_excel(changes, release_note_path, envs, env_list)
 
    result = execute(target_folder_x, envs[0],envs[1], repo_url)
    print("The result is: ", result)
 
 
 
    yaml_path = os.path.dirname(le_new_json_path)
 
    # fetch_jira_details_from_yaml_folder(github_token, repos, "dev1-lz", yaml_path)
 
    create_release_note_summary(yaml_path, target_folder_x, release_note_path, envs[1])
 
    create_upgrade_services_txt(release_note_path, envs[1], target_folder_x, envs[0])

    # configure_git_user(target_folder_x)
    stage_commit_and_push(
        target_folder_x,
        sys.argv[2],
        f'Pushing the release_note into the branch: {sys.argv[2]}',
    )
 
    target_folder_x = os.path.dirname(target_folder_x)
    target_folder_x_1 = os.path.dirname(target_folder_x_1)
 
if __name__ == '__main__':
    main()
 
 
