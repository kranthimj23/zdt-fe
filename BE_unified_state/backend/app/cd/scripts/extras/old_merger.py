import subprocess
import sys
import tempfile
import shutil
import os
import openpyxl
from openpyxl import Workbook, load_workbook
import json
import yaml
import stat
from filecmp import cmp
import datetime
from openpyxl.utils import get_column_letter

from git_helpers import (
    inject_git_token,
    is_base_branch_exists,
    clone_single_branch_and_checkout,
    clone_branch_and_checkout_new_branch,
    configure_git_user,
    stage_commit_and_push,
    stage_specific_files_commit_and_push,
)
 
 
########################################################################################################################
#### To clone the meta sheet####
########################################################################################################################
 
        
 
########################################################################################################################

def pick_branch(sheet, lower_col, branch_to_pick = None):
    for row in range(sheet.max_row, 1, -1):
        val = sheet.cell(row=row, column=lower_col).value
        if val is None or val == 'X':
            continue
        elif str(val).strip() == str(branch_to_pick).strip():
            return val, row
    raise ValueError(f"No branch found in column index {lower_col}")

def find_column_index(sheet, env_name):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == env_name:
            return col
    raise ValueError(f"Environment '{env_name}' not found in header")
 
def find_last_updated_branch(sheet, col_idx):
    for row in range(sheet.max_row, 1, -1):
        val = sheet.cell(row=row, column=col_idx).value
        if val is not None and val != 'X':
            return val, row
    raise ValueError(f"No branch found in column index {col_idx}")
 
def create_new_branch(base_branch, new_version=None):
    if new_version:
        # Assuming branch format like PF-release/1.0.0, keep prefix and append new_version
        prefix = "release"
        new_branch = f"{prefix}/{new_version}"
    else:
        new_branch = f"{base_branch}_promotion_branch"
    ##print(f"Created new branch '{new_branch}' from base branch '{base_branch}'")
    return new_branch
 
def update_excel_with_new_branch(file_path, sheet, lower_env, new_branch):
    lower_col = find_column_index(sheet, lower_env)
    max_row = sheet.max_row + 1  # next empty row
 
    # Write new branch name in lower_env column
    sheet.cell(row=max_row, column=lower_col, value=new_branch)
 
    # Mark other env columns as 'X'
    for col in range(1, sheet.max_column + 1):
        if col != lower_col:
            sheet.cell(row=max_row, column=col, value='X')
 
    # Save workbook
    sheet.parent.save(file_path)
    #print(f"Excel updated: new branch '{new_branch}' written to '{lower_env}' column at row {max_row}.")
 
def fetch_branches(file_path, lower_env, update_lower_env, new_branch_created, higher_env, promotion_branch = None, new_version = None, branch_to_pick = None):
    wb = load_workbook(filename=file_path)
    sheet = wb.active
 
    lower_col = find_column_index(sheet, lower_env)
    higher_col = find_column_index(sheet, higher_env)

    #hotfix short-circuit
    if branch_to_pick:
        new_branch_created = True
        lower_branch, _ = pick_branch(sheet, lower_col, branch_to_pick = branch_to_pick)
        higher_branch = lower_branch
        new_branch = create_new_branch(lower_branch, new_version)
        update_excel_with_new_branch(file_path, sheet, lower_env, new_branch)
        return lower_branch, new_branch, update_lower_env, new_branch_created
 
    if promotion_branch:
        lower_branch, _ = pick_branch(sheet, lower_col, branch_to_pick = promotion_branch)
    else:
        lower_branch, _ = find_last_updated_branch(sheet, lower_col)

    higher_branch, _ = find_last_updated_branch(sheet, higher_col)
 
    if lower_env == 'dev' and lower_branch!= 'X':
        if lower_branch == higher_branch:
            if not new_version:
                raise ValueError("new_version must be provided when lower_env is 'dev'")
            new_branch = create_new_branch(lower_branch, new_version)
            update_excel_with_new_branch(file_path, sheet, lower_env, new_branch)
            #print("lower branch == higher branch in dev ---> " , lower_branch, new_branch)
            new_branch_created = True
            return lower_branch, new_branch, update_lower_env, new_branch_created
 
        if lower_branch != higher_branch and higher_branch != 'X':
            #print("lower branch != higher branch in dev ---> " , higher_branch, lower_branch)
            return higher_branch, lower_branch, update_lower_env, new_branch_created
    else:
        if lower_env != 'dev' and lower_branch == higher_branch:
            new_branch = create_new_branch(lower_branch, new_version)
            update_lower_env = True
            #print(f"New branch is created values to be populated in Lower-env-->{update_lower_env}")
            update_excel_with_new_branch(file_path, sheet, 'dev', new_branch)
            return lower_branch, new_branch, update_lower_env, new_branch_created
        else:
            #print(f"Branches differ. Promoting using existing branch '{lower_branch}' in '{lower_env}'.")
            return higher_branch, lower_branch, update_lower_env, new_branch_created
 
 
########################################################################################################################
 
####
    #### To create a new branch ####
####
def clean_non_dev_folders(temp_dir):
    """Remove all files in environment folders"""
    env_folders = []
    for folder in os.listdir(temp_dir):
        item_path = os.path.join(temp_dir, f"helm-charts")
        if folder == "helm-charts":
            for i in os.listdir(item_path):
                if i.endswith('values'):    #and not i.startswith('dev'):
                    env_folders.append(i)
                    env_path = os.path.join(item_path, i)
 
                    if os.path.exists(env_path):
 
                        # Remove all files while preserving directory structure
                        for root, dirs, files in os.walk(env_path):
                            for file in files:
                                file_path = os.path.join(root, file)
                                os.remove(file_path)
                                readme_path = os.path.join(root, "readme.md")
                                with open(readme_path, 'w') as f:
                                    f.write(f"The files of {env_path} are stored here")
 
def create_github_branch(github_url, base_branch, new_branch):
    """
    Creates a new GitHub branch from an existing branch and updates meta-sheet.xlsx
    - Clears files in all *values folders except dev-values
    - Preserves directory structure
    - Updates meta-sheet in master branch
    """
    temp_dir = tempfile.mkdtemp()
    try:
        if not is_base_branch_exists(github_url, base_branch):
            raise ValueError(f"Base branch '{base_branch}' not found in repository")

        new_branch_dir = os.path.join(temp_dir, "new_branch")
        clone_branch_and_checkout_new_branch(github_url, base_branch, new_branch, new_branch_dir)

        clean_non_dev_folders(new_branch_dir)

        configure_git_user(new_branch_dir)
        stage_commit_and_push(
            new_branch_dir,
            new_branch,
            f'Initialize {new_branch}: Clean environment folders',
            pull_before_push=False,
        )

        master_dir = os.path.join(temp_dir, "master")
        clone_single_branch_and_checkout(github_url, 'master', master_dir)

        excel_path = os.path.join(master_dir, 'meta-sheet.xlsx')
        wb = load_workbook(excel_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        dev_col = headers.index('dev') + 1

        next_row = 2
        while ws.cell(row=next_row, column=dev_col).value:
            next_row += 1

        ws.cell(row=next_row, column=dev_col).value = new_branch

        for col in range(1, ws.max_column + 1):
            if col == dev_col:
                continue
            cell = ws.cell(row=next_row, column=col)
            if not cell.value:
                cell.value = 'X'

        wb.save(excel_path)

        configure_git_user(master_dir)
        stage_specific_files_commit_and_push(
            master_dir,
            'master',
            f'Add {new_branch} to meta-sheet',
            ['meta-sheet.xlsx'],
            pull_before_push=False,
        )

    except subprocess.CalledProcessError as e:
        return False
    except Exception as e:
        return False
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    return True
 
########################################################################################################################
 
def main():
    repo_name = "promotion-repo"
    lower_env = sys.argv[1]                  # Change as needed
    higher_env = sys.argv[2]
    github_url = sys.argv[3]                # Change as needed
    new_version = sys.argv[4]              # Provide new version for dev branch creation
    branch_to_pick = sys.argv[5] if len(sys.argv) > 5 else None
    target_folder = os.path.join(os.getcwd(), 'execution')
    os.makedirs(target_folder, exist_ok=True)
    
    github_url = inject_git_token(github_url)

    meta_sheet_file_path = os.path.join(target_folder, f"meta-sheet.xlsx")
    update_lower_env = False
    reverse_promotion = False
    new_branch_created = False
    clone_single_branch_and_checkout(github_url, "master", target_folder)

    if lower_env != 'dev1' and lower_env != 'dev2':
        promotion_branch = f"release-{new_version}"
        prev_branch, present_branch, update_lower_env, new_branch_created = fetch_branches(meta_sheet_file_path, lower_env, update_lower_env, new_branch_created, higher_env, promotion_branch = promotion_branch, new_version = new_version)
    else:
        prev_branch, present_branch, update_lower_env, new_branch_created = fetch_branches(meta_sheet_file_path, lower_env, update_lower_env, new_branch_created, higher_env, new_version = new_version, branch_to_pick=branch_to_pick)

    if new_branch_created:
        create_github_branch(github_url,prev_branch,present_branch, lower_env)
    # prev_branch, present_branch, update_lower_env, new_branch_created = fetch_branches(meta_sheet_file_path, lower_env, update_lower_env, new_branch_created,higher_env, new_version)
    # #print(f"Current branch in '{lower_env}': {prev_branch}")
    # if new_branch_created:
    #     create_github_branch(github_url,prev_branch,present_branch)
    #     #print("new branch will be created")
 
    #print(prev_branch, present_branch, update_lower_env, new_branch_created)
    # else:
    #     #print("No new branch created; promoting using existing branch.")
    envs = []
 
    if update_lower_env:
        envs.append('dev')
        # envs.append(lower_env)
        envs.append(lower_env)
        reverse_promotion = True
    else:
        envs.append(lower_env)
        envs.append(higher_env)
 
    def on_rm_error(func, path, exc_info):
     os.chmod(path, stat.S_IWRITE)
     func(path)
 
    shutil.rmtree('execution', onerror=on_rm_error)
    x1 = prev_branch
    x2 = present_branch
    low = envs[0]
    high = envs[1]
    isNew = new_branch_created
    print(f"{x1}, {x2}, {low}, {high}, {isNew}")
 
if __name__ == "__main__":
    main()
########################################################################################################################