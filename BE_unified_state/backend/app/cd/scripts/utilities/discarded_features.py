import os
import openpyxl
import shutil
import re
import yaml
from openpyxl.utils import get_column_letter

def create_release_note_summary(files_path, target_folder_x, existing_release_note_dir,input_sheet_name):
    print("Executing create release note.py")
    # Find first Excel file in existing release note folder
    excel_file = None
    if os.path.exists(existing_release_note_dir):
        for f in os.listdir(existing_release_note_dir):
            if f.endswith(".xlsx"):
                excel_file = os.path.join(existing_release_note_dir, f)
                break
 
    if not excel_file:
        print("Excel sheet not found in the source release note directory.")
        return
 
    # Ensure new release note directory exists
    if not os.path.exists(target_folder_x):
        print("Path to summary sheet not specified")
 
    # Compose the directory in which the file is to be copied
    new_dir = os.path.join(target_folder_x, "helm-charts", f"{input_sheet_name}-values", "release_note")
 
    # Create directory if it doesn't exist
    os.makedirs(new_dir, exist_ok=True)
 
    # Then, compose the full new Excel file path
    new_excel_file = os.path.join(new_dir, "release-note-summary.xlsx")
    # Copy original Excel file to this new location
    try:
        shutil.copy2(excel_file, new_excel_file)
        print(f"Copied {excel_file} to {new_excel_file}")
    except Exception as e:
        print(f"Failed to copy Excel file: {e}")
        return
 
    # Continue with rest of your processing by loading new_excel_file...
    wb = openpyxl.load_workbook(new_excel_file)
 
    # Get YAML files from the specified path
    yaml_files = [f for f in os.listdir(files_path) if f.endswith('.yaml') or f.endswith('.yml')]
 
    # Check if "Summary" sheet already exists; if so, remove it
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
 
    # Create a new "Summary" sheet
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.sheet_state = "visible"
 
    # Add column headers including new "Build Number" at third position
    headers = ["Services", "Status", "Build Number", "Comments", "Owner"]
    summary_sheet.append(headers)
 
    # Add service names from YAML files
    for yaml_file in yaml_files:
        service_name = os.path.splitext(yaml_file)[0]  # Use file name (without extension) as service name
        # Initialize row with empty values for Status, Build Number, Comments, Owner
        summary_sheet.append([service_name, "", "", "", ""])
 
    # Add dropdown options to the "Status" column
    dropdown_options = ["Updated", "No modifications", "New service", "Deleted service"]
 
    # Data validation for dropdowns
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"{}"'.format(",".join(dropdown_options)),
        allow_blank=True
    )
 
    # Add the dropdown to all cells in the "Status" column (from row 2 onwards)
    for row in range(2, summary_sheet.max_row + 1):
        cell = summary_sheet.cell(row=row, column=2)  # Status column
        dv.add(cell.coordinate)
 
    summary_sheet.add_data_validation(dv)
 
    # Load the specified sheet for comments
    if input_sheet_name in wb.sheetnames:
        comments_sheet = wb[input_sheet_name]
    else:
        print(f"Sheet '{input_sheet_name}' not found in the Excel file.")
        return
 
    # Function to extract build number from image_name string
    def extract_build_number(image_name_value):
        # Example: format x.y.z-bpqr-qwertyu-0.0
        # Goal: find substring starting with 'b' (like 'bpqr') and extract numeric portion (pqr)
        # This regex captures the numeric part following 'b' in the dash-separated substring after the last hyphen
        # but from your example, the number is after b in bpqr (where pqr is digits)
        # We'll extract the substring starting with b and then digits after it (with some leniency)
        # If digits can be anywhere after 'b', this regex helps extract first digits after b
 
        # A more concrete approach:
        # Find segment after first 'b' in the string where digits appear consecutively
 
        match = re.search(r"-b(\d+)", image_name_value) # '-' followed by 'b' and digits
        if match:
            return match.group(1)  # return digits after 'b'
        else:
            # Alternatively, more relaxed pattern: find 'b' followed by digits anywhere
            match = re.search(r"b(\d+)", image_name_value)
            if match:
                return match.group(1)
        return ""  # Return empty string if no match
 
    # Create a dictionary for comments to speed up lookup: {service_name: comment}
    comments_dict = {}
    for r in range(2, comments_sheet.max_row + 1):
        svc_name = comments_sheet.cell(row=r, column=1).value
        comment = comments_sheet.cell(row=r, column=comments_sheet.max_column).value  # Assuming last column has comment
        if svc_name:
            comments_dict[svc_name] = comment
 
    # Iterate through rows in Summary sheet to update Status and Build Number
    for row in range(2, summary_sheet.max_row + 1):
        service_name = summary_sheet.cell(row=row, column=1).value
 
        comment = comments_dict.get(service_name, None)
        # Update status based on comment
        status = "No modifications"
        if comment == "root object added":
            status = "New service"
        elif comment == "root object deleted":
            status = "Deleted service"
        elif comment in ["Modified", "Added"]:
            status = "Updated"
 
        summary_sheet.cell(row=row, column=2).value = status
 
        # Build Number column is at column 3 (index 3)
        build_number_cell = summary_sheet.cell(row=row, column=3)
 
        # If status is Updated or New service, parse YAML and extract build number
        if status in ["Updated", "New service"]:
            # Compose full path to yaml file
            yaml_file_path = os.path.join(files_path, f"{service_name}.yaml")
            if not os.path.exists(yaml_file_path):
                # Try .yml extension if .yaml does not exist
                yaml_file_path = os.path.join(files_path, f"{service_name}.yml")
 
            if os.path.exists(yaml_file_path):
                try:
                    with open(yaml_file_path, 'r') as yf:
                        yaml_content = yaml.safe_load(yf)
 
                    # Search for 'image_name' key anywhere in yaml (top level or nested)
                    # If your structure is nested, you may need a recursive search
                    def find_image_name(data):
                        if isinstance(data, dict):
                            for k, v in data.items():
                                if k == "image_name":
                                    return v
                                else:
                                    found = find_image_name(v)
                                    if found:
                                        return found
                        elif isinstance(data, list):
                            for item in data:
                                found = find_image_name(item)
                                if found:
                                    return found
                        return None
 
                    image_name_val = find_image_name(yaml_content)
                    if image_name_val:
                        build_num = extract_build_number(str(image_name_val))
                        build_number_cell.value = build_num
                    else:
                        build_number_cell.value = ""
 
                except Exception as e:
                    print(f"Failed to read/parse {yaml_file_path}: {e}")
                    build_number_cell.value = ""
            else:
                build_number_cell.value = ""
        else:
            build_number_cell.value = ""
 
    # Adjust column widths for better readability (columns A to E now)
    for col in range(1, 6):  # A=1 to E=5
        column_letter = get_column_letter(col)
        summary_sheet.column_dimensions[column_letter].width = 20
 
    # Save the workbook with the updated Summary sheet
    wb.save(new_excel_file)
 
    # Re-open and reorder sheets to put Summary first
    wb = openpyxl.load_workbook(new_excel_file)
 
    if "Summary" in wb.sheetnames:
        sheet = wb["Summary"]
 
        # Get the current index of the sheet
        current_index = wb._sheets.index(sheet)
 
        # If the sheet is not already the first one, move it
        if current_index != 0:
            # Remove the sheet from its current position
            wb._sheets.pop(current_index)
            # Insert the sheet at the beginning
            wb._sheets.insert(0, sheet)
 
        wb.save(new_excel_file)