
import os
import shutil

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from utilities.helpers import extract_hyperlink_path


def pick_branch(sheet, lower_col, branch_to_pick=None):
    if not branch_to_pick:
        raise ValueError("branch_to_pick required")
    
    for row in range(sheet.max_row, 0, -1):  # Include row 1 if needed
        val = sheet.cell(row=row, column=lower_col).value
        if val and str(val).strip() == str(branch_to_pick).strip():
            return val, row
    raise ValueError(f"Branch '{branch_to_pick}' not in col {lower_col}")


def get_sheet(sheet_path):
    return load_workbook(filename=sheet_path).active


def get_cell_value(row_num, col_name, headers, ws): 
    col_idx = headers.get(col_name)
    if col_idx is None:
        return None
    return ws.cell(row=row_num, column=col_idx).value


def get_headers(headers, ws):
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[cell.value] = idx
    return headers


def find_column_index(sheet, env_name):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == env_name:
            return col
    raise ValueError(f"Environment '{env_name}' not found in header")


def create_new_branch(new_version=None):
    if not new_version:
                raise ValueError("new_version must be provided when lower_env is 'dev1'")
    return f"release/{new_version}"


def find_last_updated_branch(sheet, col_idx):
    for row in range(sheet.max_row, 1, -1):
        val = sheet.cell(row=row, column=col_idx).value
        if val is not None and val != 'X':
            return val, row
    raise ValueError(f"No branch found in column index {col_idx}")


def find_promotion_branch(sheet, lower_col, promotion_branch):
    for row in range(sheet.max_row, 1, -1):
        val = sheet.cell(row=row, column=lower_col).value
        if val is None or val == 'X':
            continue
        if str(val).strip() == str(promotion_branch).strip():
            return val, row
    raise ValueError(f"No branch found in column index {lower_col}")


def find_branch(sheet, col_idx, promoting = None, target = None):
    for row in range(sheet.max_row, 1, -1):
        val = sheet.cell(row=row, column=col_idx).value
        if val is None or val == 'X':
            continue
        elif promoting:
            if str(val).strip() == str(target).strip():
                return val, row
        else:
            return val, row
    col_name = get_column_letter(col_idx)
    raise ValueError(f"No branch found in column {col_name}")


def create_new_drop_in_meta_sheet(file_path, sheet, lower_env, new_branch):
    lower_col = find_column_index(sheet, lower_env)
    max_row = sheet.max_row + 1
    sheet.cell(row=max_row, column=lower_col, value=new_branch)
    for col in range(1, sheet.max_column + 1):
        if col != lower_col:
            sheet.cell(row=max_row, column=col, value='X')
    sheet.parent.save(file_path)


def get_sheets_with_values(excel_file, env):
    wb = load_workbook(excel_file, data_only=True)
    matching_sheets = ''
 
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Summary", "AQL", "SQL"):
            print("Skipping the sheet: ", sheet_name)
        else:
            ws = wb[sheet_name]
 
            # Find the column index of the "env-current value" header
            value_col_idx = None
            for cell in ws[1]:
                if cell.value == f"{env}-current value":
                    value_col_idx = cell.column  # 1-based index
                    break
 
            if value_col_idx is not None:
                for row in ws.iter_rows(min_row=2, min_col=value_col_idx, max_col=value_col_idx + 2):
                    cell_value = row[0].value
                    comment = row[2].value
                    if comment == "Deleted" or cell_value is not None and str(cell_value).strip() != "":
                        matching_sheets = sheet_name
                        break  # No need to check further cells in this sheet

    if not matching_sheets:
        print("No values updated in any sheet for promotion.")
    else:
        print("Sheets with values:", matching_sheets)
 
    return matching_sheets