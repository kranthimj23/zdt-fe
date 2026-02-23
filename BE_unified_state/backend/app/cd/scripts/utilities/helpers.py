import re
import os
import json
import shutil
import urllib.parse
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

def tokenize_url(git_repo_url):
    #username - password for debugging
    github_username = "binaryBodhi"
    # github_password = "placeholder"
    # github_username = os.getenv("GIT_USER")
    github_password = os.getenv("GIT_TOKEN")
    # encoded_password = urllib.parse.quote(github_password, safe="")
    if git_repo_url.startswith("https://github"):
        git_repo_url = git_repo_url.replace("https://", f"https://{github_username}:{github_password}@")
        return git_repo_url
    raise ValueError("Unsupported repository_url format. Must start with https://")


def extract_hyperlink_path(hyperlink_formula):
    """
    Extracts the file path from an Excel HYPERLINK formula.
    Assumes the formula is in the format: =HYPERLINK("path_to_file", "text")
    """
    # Extract the part inside the parentheses
    match = re.search(r'file:////(.*)', hyperlink_formula)
    if match:
        return match.group(1)
    return None

def get_parent_path(path: str) -> str:
    if not path:
        return path
    return path.rsplit("//", 1)[0] if "//" in path else path

def create_upgrade_services_txt(excel_path, sheet_name, repo_root, lower_env):
    txt_path = os.path.join(repo_root, 'upgrade-services.txt')
    
    if os.path.exists(txt_path):
        os.remove(txt_path)
 
    # Identify the correct .xlsx file
    file_path = None
    for i in os.listdir(excel_path):
        if i.endswith(".xlsx"):
            file_path = os.path.join(excel_path, i)
            break
 
    if not file_path:
        print("No Excel file found in the given path.")
    else:
        wb = load_workbook(file_path)
 
        if sheet_name not in wb.sheetnames:
            print("Application release note does not exist")
        else:
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            try:
                key_col = headers.index('Key') + 1
                service_name_col = headers.index('Service name') + 1
                value_col = headers.index(f'{lower_env}-current value') + 1
                comment_col = headers.index(f'Comment') + 1
            except ValueError as e:
                raise ValueError(f"Required column missing: {e}")
 
            with open(txt_path, 'w') as file:
                for row in ws.iter_rows(min_row=2):
                    key_cell = row[key_col - 1].value
                    service_name = row[service_name_col - 1].value
                    value = row[value_col - 1].value
                    comment = row[comment_col - 1].value
 
                    if key_cell and 'image//image_name' in str(key_cell):
                        if service_name and value:
                            cleanValue = value.strip('"')
                            file.write(f"{service_name}:{cleanValue}\n")
                    elif comment and comment.strip().lower() == "root object added" and service_name not in ("data", "env"):
                        if value:
                            if isinstance(value, str) and value.startswith('file'):
                                hyperlink_path = rf'/{extract_hyperlink_path(value)}'
                                with open(hyperlink_path, 'r') as txt_file:
                                        value = txt_file.read().strip()
                            try:
                                parsed = json.loads(value)
                                tag = parsed.get("image", {}).get("image_name")
                                if service_name and tag:
                                    file.write(f"{service_name}:{tag}\n")
                            except json.JSONDecodeError as e:
                                print(f"Failed to parse JSON for service {service_name}: {e}")
                        else:
                            print(f"Value is None for service {service_name} with 'root object added'")
                    else:
                        print(f"Skipping service {service_name}")

    temp_path_app = os.path.join(repo_root, "helm-charts", f"{sheet_name}-values", "app-values", "deployment")
    if not os.path.exists(temp_path_app):
        os.makedirs(temp_path_app, exist_ok=True)
    file_path = os.path.join(temp_path_app, 'upgrade-services.txt')
    shutil.copy2(txt_path, file_path)